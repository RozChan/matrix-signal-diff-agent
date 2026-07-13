"""Generate signal-level AI-assisted review and human audit details."""

from __future__ import annotations

import json
import re
import time
from pathlib import Path
from typing import Any

from .llm_client import LLMConfigurationError, LLMRequestError, LLMTimeoutError, call_chat_json, get_llm_config

SOURCE_SHEETS = ["完全同名匹配对比结果", "vcu-hcu 同名匹配"]
AI_REVIEW_SHEET = "AI辅助复核与人工审核明细"
DIFF_LIST_HEADER = "差异点list"
SIGNAL_40_HEADER = "4.0信号名"
SIGNAL_51_HEADER = "5.1信号名"

TEXT_REVIEW_FIELDS = {"信号值描述", "单位"}
NUMERIC_FIELDS = {"信号长度", "精度", "偏移量", "物理最小值", "物理最大值"}
KNOWN_DIFF_FIELDS = ["信号长度", "精度", "偏移量", "物理最小值", "物理最大值", "单位", "信号值描述"]

REVIEW_HEADERS = [
    "来源Sheet",
    "4.0信号名",
    "5.1信号名",
    "差异字段汇总",
    "差异字段数量",
    "是否包含数值类差异",
    "是否包含文本类差异",
    "原始差异点list",
    "字段差异明细",
    "信号级AI判断结果",
    "差异类型汇总",
    "置信度",
    "信号级AI判断理由",
    "信号级AI建议处理方式",
    "需人工优先确认",
    "系统默认结论",
    "系统默认原因",
    "人工审核结果",
    "人工备注",
]

ALLOWED_SIGNAL_JUDGEMENTS = {"真实差异", "疑似可忽略", "无法判断"}
ALLOWED_TYPE_SUMMARIES = {"数值定义变化", "错别字", "缩写差异", "表达方式不同", "语义相近", "枚举描述写法不同", "无法判断"}
ALLOWED_CONFIDENCE = {"高", "中", "低", "不适用"}
ALLOWED_ACTIONS = {"应保留差异", "建议人工确认"}

EMPTY_STATS: dict[str, int] = {
    "total_review_items": 0,
    "signal_review_item_count": 0,
    "diff_field_count": 0,
    "text_signal_count": 0,
    "numeric_signal_count": 0,
    "ai_reviewed_count": 0,
    "ai_skipped_count": 0,
    "real_diff_count": 0,
    "suspected_ignore_count": 0,
    "unknown_count": 0,
    "llm_disabled_count": 0,
    "ai_called_count": 0,
    "ai_failed_count": 0,
}


def _header_map(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def _cell(ws, row_idx: int, headers: dict[str, int], name: str) -> str:
    col = headers.get(name)
    if not col:
        return ""
    value = ws.cell(row=row_idx, column=col).value
    return "" if value is None else str(value).strip()


def parse_diff_list(diff_text: Any) -> list[dict[str, str]]:
    """Parse 差异点list into field-level diffs for one signal."""

    text = "" if diff_text is None else str(diff_text).strip()
    if not text:
        return []

    fields_pattern = "|".join(re.escape(field) for field in KNOWN_DIFF_FIELDS)
    pattern = re.compile(
        rf"(?P<field>{fields_pattern})：4\.0=(?P<v40>.*?)；5\.1=(?P<v51>.*?)(?=\n(?:{fields_pattern})：4\.0=|\Z)",
        re.S,
    )
    items = [
        {
            "diff_field": match.group("field").strip(),
            "value_40": match.group("v40").strip(),
            "value_51": match.group("v51").strip(),
            "field_type": field_type(match.group("field").strip()),
        }
        for match in pattern.finditer(text)
    ]
    if items:
        return items
    return [{"diff_field": "未解析", "value_40": "", "value_51": "", "field_type": "unknown"}]


def field_type(diff_field: str) -> str:
    if diff_field in NUMERIC_FIELDS:
        return "numeric"
    if diff_field in TEXT_REVIEW_FIELDS:
        return "text"
    return "unknown"


def field_type_label(value: str) -> str:
    return {"numeric": "数值类差异", "text": "文本类差异", "unknown": "未解析"}.get(value, "未解析")


def _format_field_diffs(field_diffs: list[dict[str, str]]) -> str:
    blocks = []
    for diff in field_diffs:
        blocks.append(
            f"【{diff.get('diff_field', '')}】\n"
            f"4.0：{diff.get('value_40', '')}\n"
            f"5.1：{diff.get('value_51', '')}\n"
            f"类型：{field_type_label(diff.get('field_type', 'unknown'))}"
        )
    return "\n\n".join(blocks)


def _build_signal_item(source_sheet: str, signal_40: str, signal_51: str, diff_text: str) -> dict[str, Any]:
    field_diffs = parse_diff_list(diff_text)
    diff_fields = [diff["diff_field"] for diff in field_diffs]
    has_numeric = any(diff.get("field_type") == "numeric" for diff in field_diffs)
    has_text = any(diff.get("field_type") == "text" for diff in field_diffs)
    return {
        "source_sheet": source_sheet,
        "signal_40": signal_40,
        "signal_51": signal_51,
        "diff_fields": diff_fields,
        "diff_field_count": len(field_diffs),
        "has_numeric_diff": has_numeric,
        "has_text_diff": has_text,
        "original_diff_list": diff_text,
        "field_diffs": field_diffs,
        "field_diff_details": _format_field_diffs(field_diffs),
    }


def _iter_signal_items(wb) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for source_sheet in SOURCE_SHEETS:
        if source_sheet not in wb.sheetnames:
            continue
        ws = wb[source_sheet]
        headers = _header_map(ws)
        for row_idx in range(2, (ws.max_row or 1) + 1):
            diff_text = _cell(ws, row_idx, headers, DIFF_LIST_HEADER)
            if not diff_text:
                continue
            rows.append(
                _build_signal_item(
                    source_sheet,
                    _cell(ws, row_idx, headers, SIGNAL_40_HEADER),
                    _cell(ws, row_idx, headers, SIGNAL_51_HEADER),
                    diff_text,
                )
            )
    return rows


def _safe_choice(value: Any, allowed: set[str], default: str) -> str:
    text = "" if value is None else str(value).strip()
    return text if text in allowed else default


def _prompt_messages(item: dict[str, Any]) -> list[dict[str, str]]:
    system = (
        "你是车辆信号矩阵差异复核助手。你会收到同一个信号在 4.0 和 5.1 中的所有差异字段。"
        "请基于全部差异字段判断该信号整体是否属于：真实差异、疑似可忽略、无法判断。"
        "判断原则：1. 如果存在信号长度、精度、偏移量、物理最小值、物理最大值等数值类定义差异，应判断为真实差异。"
        "2. 只有当差异仅为单位或信号值描述等文本类差异，并且这些差异可能只是错别字、缩写、表达方式不同、语义相近或枚举写法不同，才可以判断为疑似可忽略。"
        "3. 如果无法仅根据输入内容判断，请输出无法判断。4. 不允许编造信号业务含义。"
        "5. 不允许把明显不同的信号定义强行判断为可忽略。6. AI 结果只作为人工审核参考。"
        "输出必须是 JSON，不要输出 Markdown。"
    )
    user_payload = {
        "source_sheet": item["source_sheet"],
        "signal_40": item["signal_40"],
        "signal_51": item["signal_51"],
        "original_diff_list": item["original_diff_list"],
        "diffs": item["field_diffs"],
        "task": "请基于该信号的全部差异字段，判断这个信号整体是否属于真实差异、疑似可忽略或无法判断。",
        "输出 JSON 格式": {
            "signal_ai_judgement": "真实差异/疑似可忽略/无法判断",
            "difference_type_summary": "数值定义变化/错别字/缩写差异/表达方式不同/语义相近/枚举描述写法不同/无法判断",
            "confidence": "高/中/低",
            "signal_ai_reason": "简短理由",
            "signal_ai_suggested_action": "应保留差异/建议人工确认",
        },
    }
    return [
        {"role": "system", "content": system},
        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
    ]


def _system_real_diff_review(item: dict[str, Any]) -> dict[str, str]:
    numeric_fields = [diff["diff_field"] for diff in item["field_diffs"] if diff.get("field_type") == "numeric"]
    fields_text = "、".join(numeric_fields) or "数值类定义字段"
    return {
        "ai_reviewed": "否",
        "signal_ai_judgement": "真实差异",
        "difference_type_summary": "数值定义变化",
        "confidence": "高",
        "signal_ai_reason": f"该信号存在{fields_text}等数值类定义差异，系统判定为真实差异。",
        "signal_ai_suggested_action": "应保留差异",
    }


def _disabled_review() -> dict[str, str]:
    return {
        "ai_reviewed": "否",
        "signal_ai_judgement": "未启用",
        "difference_type_summary": "无法判断",
        "confidence": "不适用",
        "signal_ai_reason": "AI辅助复核未启用，文本类差异需人工确认。",
        "signal_ai_suggested_action": "建议人工确认",
    }


def _unknown_review(reason: str) -> dict[str, str]:
    return {
        "ai_reviewed": "否",
        "signal_ai_judgement": "无法判断",
        "difference_type_summary": "无法判断",
        "confidence": "低",
        "signal_ai_reason": reason,
        "signal_ai_suggested_action": "建议人工确认",
    }


def _ai_review(item: dict[str, Any]) -> dict[str, str]:
    data = call_chat_json(_prompt_messages(item))
    judgement = _safe_choice(data.get("signal_ai_judgement"), ALLOWED_SIGNAL_JUDGEMENTS, "无法判断")
    action_default = "应保留差异" if judgement == "真实差异" else "建议人工确认"
    return {
        "ai_reviewed": "是",
        "signal_ai_judgement": judgement,
        "difference_type_summary": _safe_choice(data.get("difference_type_summary"), ALLOWED_TYPE_SUMMARIES, "无法判断"),
        "confidence": _safe_choice(data.get("confidence"), ALLOWED_CONFIDENCE - {"不适用"}, "低"),
        "signal_ai_reason": str(data.get("signal_ai_reason") or "模型未返回理由").strip(),
        "signal_ai_suggested_action": _safe_choice(data.get("signal_ai_suggested_action"), ALLOWED_ACTIONS, action_default),
    }


def get_signal_level_ai_judgement(item: dict[str, Any], enable_ai: bool, llm_enabled: bool, stats: dict[str, Any]) -> dict[str, str]:
    if item.get("has_numeric_diff"):
        return _system_real_diff_review(item)
    if not item.get("has_text_diff"):
        return _unknown_review("差异字段未解析或不属于可复核文本字段，需人工确认。")
    if not enable_ai or not llm_enabled:
        return _disabled_review()
    stats["ai_called_count"] += 1
    try:
        return _ai_review(item)
    except LLMTimeoutError as exc:
        stats["ai_failed_count"] += 1
        warning = str(exc)
        if warning not in stats["warnings"]:
            stats["warnings"].append(warning)
        return _unknown_review("模型请求超时，需人工确认。")
    except (LLMConfigurationError, LLMRequestError) as exc:
        stats["ai_failed_count"] += 1
        warning = str(exc)
        if warning not in stats["warnings"]:
            stats["warnings"].append(warning)
        return _unknown_review(f"AI辅助复核失败：{warning}")


def _default_review_result(review: dict[str, str]) -> tuple[str, str, str]:
    if review["signal_ai_judgement"] == "真实差异":
        return "确认真实差异", "AI或规则判断该信号存在真实差异，系统默认保留；人工可修改", "否"
    if review["signal_ai_judgement"] == "疑似可忽略":
        return "", "AI判断该信号差异疑似可忽略，需人工优先确认", "是"
    return "", "AI未给出可直接采用的结论，需人工确认", "否"


def _stats_increment(stats: dict[str, Any], item: dict[str, Any], review: dict[str, str]) -> None:
    stats["total_review_items"] += 1
    stats["signal_review_item_count"] += 1
    stats["diff_field_count"] += int(item.get("diff_field_count") or 0)
    if item.get("has_numeric_diff"):
        stats["numeric_signal_count"] += 1
    if item.get("has_text_diff"):
        stats["text_signal_count"] += 1
    if review["ai_reviewed"] == "是":
        stats["ai_reviewed_count"] += 1
    else:
        stats["ai_skipped_count"] += 1
    if review["signal_ai_judgement"] == "真实差异":
        stats["real_diff_count"] += 1
    elif review["signal_ai_judgement"] == "疑似可忽略":
        stats["suspected_ignore_count"] += 1
    elif review["signal_ai_judgement"] == "无法判断":
        stats["unknown_count"] += 1
    elif review["signal_ai_judgement"] == "未启用":
        stats["llm_disabled_count"] += 1


def _style_sheet(ws) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    header_fill = PatternFill("solid", fgColor="7030A0")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {
        "A": 24, "B": 36, "C": 36, "D": 28, "E": 14, "F": 16, "G": 16, "H": 70, "I": 80,
        "J": 18, "K": 20, "L": 12, "M": 55, "N": 20, "O": 16, "P": 18, "Q": 55, "R": 18, "S": 40,
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def run_ai_review(compare_file_path: Path, enable_ai: bool = False, progress_callback=None) -> dict[str, Any]:
    """Append signal-level AI辅助复核与人工审核明细 sheet to the compare workbook."""

    compare_path = Path(compare_file_path)
    started = time.perf_counter()
    stats: dict[str, Any] = {**EMPTY_STATS, "warnings": []}
    config = get_llm_config()
    llm_enabled = config.enabled

    def emit(**payload):
        if progress_callback is not None:
            progress_callback(payload)

    from openpyxl import load_workbook

    emit(stage="正在解析信号级差异明细")
    wb = load_workbook(compare_path)
    if AI_REVIEW_SHEET in wb.sheetnames:
        del wb[AI_REVIEW_SHEET]
    signal_items = _iter_signal_items(wb)
    total_signals = len(signal_items)
    total_fields = sum(int(item.get("diff_field_count") or 0) for item in signal_items)
    ai_required_signals = sum(1 for item in signal_items if not item.get("has_numeric_diff") and item.get("has_text_diff"))
    system_direct_signals = total_signals - ai_required_signals
    if not enable_ai or not llm_enabled:
        emit(
            stage="AI 未启用，仅生成信号级人工审核清单",
            total=total_signals,
            field_total=total_fields,
            ai_required_total=0,
            system_direct_total=system_direct_signals,
        )
    else:
        emit(
            stage="正在执行信号级 AI 复核",
            total=total_signals,
            field_total=total_fields,
            ai_required_total=ai_required_signals,
            system_direct_total=system_direct_signals,
        )

    ws = wb.create_sheet(AI_REVIEW_SHEET)
    ws.append(REVIEW_HEADERS)

    completed = 0
    ai_completed = 0
    for index, item in enumerate(signal_items, start=1):
        emit(
            stage="正在执行信号级 AI 复核" if enable_ai and llm_enabled else "AI 未启用，仅生成信号级人工审核清单",
            current=index,
            total=total_signals,
            field_total=total_fields,
            ai_required_total=ai_required_signals if enable_ai and llm_enabled else 0,
            system_direct_total=system_direct_signals,
            signal_name=item.get("signal_40") or item.get("signal_51") or "",
            completed=completed,
            ai_completed=ai_completed,
            failed=stats["ai_failed_count"],
        )
        review = get_signal_level_ai_judgement(item, enable_ai, llm_enabled, stats)
        _stats_increment(stats, item, review)
        if review.get("ai_reviewed") == "是" or (enable_ai and llm_enabled and not item.get("has_numeric_diff") and item.get("has_text_diff")):
            ai_completed += 1
        completed += 1
        default_result, default_reason, priority = _default_review_result(review)
        ws.append([
            item["source_sheet"],
            item["signal_40"],
            item["signal_51"],
            "、".join(item["diff_fields"]),
            item["diff_field_count"],
            "是" if item["has_numeric_diff"] else "否",
            "是" if item["has_text_diff"] else "否",
            item["original_diff_list"],
            item["field_diff_details"],
            review["signal_ai_judgement"],
            review["difference_type_summary"],
            review["confidence"],
            review["signal_ai_reason"],
            review["signal_ai_suggested_action"],
            priority,
            default_result,
            default_reason,
            "",
            "",
        ])

    emit(
        stage="正在写入 Excel",
        total=total_signals,
        field_total=total_fields,
        ai_required_total=ai_required_signals if enable_ai and llm_enabled else 0,
        system_direct_total=system_direct_signals,
        completed=completed,
        ai_completed=ai_completed,
        failed=stats["ai_failed_count"],
    )
    _style_sheet(ws)
    wb.save(compare_path)
    wb.close()
    stats["elapsed_seconds"] = round(time.perf_counter() - started, 2)
    return stats
