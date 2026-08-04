"""Idempotent delivery of the three final Excel files to a Feishu group."""

from __future__ import annotations

import os
import re
import shutil
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .ai_review import AI_REVIEW_SHEET, SOURCE_SHEETS
from .feishu_openapi_client import FeishuOpenAPIClient
from .final_export import FINAL_REVIEW_FILENAME, FINAL_RESULT_HEADERS
from .lark_cli_client import LarkCliClient
from .pipeline import OUTPUT_FILENAMES
from .review_store import load_task_meta, update_task_meta, utc_now_iso
from .task_lock import get_task_lock


DELIVERY_ORDER = ("full_40", "full_51", "compare_final")
_BASELINE_RE = re.compile(r"(?i)(?<![A-Z0-9])(\d{2}R\d+)(?![A-Z0-9])")


class FeishuFileDeliveryError(RuntimeError):
    def __init__(self, stage: str, message: str) -> None:
        self.stage = stage
        super().__init__(message)


def _enabled() -> bool:
    return os.getenv("FEISHU_RESULT_FILE_DELIVERY_ENABLED", "true").strip().lower() == "true"


def register_final_result_files(task_dir: Path, final_path: Path) -> dict[str, str]:
    tdir = Path(task_dir).resolve()
    output = tdir / "output"
    expected_final = output / FINAL_REVIEW_FILENAME
    if Path(final_path).resolve() != expected_final.resolve():
        raise FeishuFileDeliveryError("validate", "第三个附件不是当前任务的最终人工审核结果")
    registered = {
        "full_40": str((Path("output") / OUTPUT_FILENAMES["full_40"])),
        "full_51": str((Path("output") / OUTPUT_FILENAMES["full_51"])),
        "compare_final": str((Path("output") / FINAL_REVIEW_FILENAME)),
    }
    update_task_meta(tdir, final_result_files=registered)
    return registered


def _registered_files(task_dir: Path, meta: dict[str, Any]) -> dict[str, Path]:
    tdir = Path(task_dir).resolve()
    expected = {
        "full_40": Path("output") / OUTPUT_FILENAMES["full_40"],
        "full_51": Path("output") / OUTPUT_FILENAMES["full_51"],
        "compare_final": Path("output") / FINAL_REVIEW_FILENAME,
    }
    registered = dict(meta.get("final_result_files") or {})
    paths: dict[str, Path] = {}
    for key in DELIVERY_ORDER:
        if Path(str(registered.get(key) or "")) != expected[key]:
            raise FeishuFileDeliveryError("validate", f"任务未登记正确的结果文件：{key}")
        path = (tdir / expected[key]).resolve()
        if path.parent != (tdir / "output").resolve():
            raise FeishuFileDeliveryError("validate", "结果文件路径越界")
        if not path.is_file():
            raise FeishuFileDeliveryError("validate", f"结果文件不存在：{path.name}")
        if path.stat().st_size <= 0:
            raise FeishuFileDeliveryError("validate", f"结果文件为空：{path.name}")
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise FeishuFileDeliveryError("validate", f"结果文件不是有效Excel：{path.name}") from exc
        try:
            if key == "compare_final":
                if workbook.sheetnames != list(SOURCE_SHEETS):
                    raise FeishuFileDeliveryError("validate", "最终差异文件必须且只能保留两个原始差异Sheet")
                if AI_REVIEW_SHEET in workbook.sheetnames:
                    raise FeishuFileDeliveryError("validate", "最终差异文件仍包含AI辅助复核明细Sheet")
                for sheet_name in SOURCE_SHEETS:
                    headers = {str(cell.value or "").strip() for cell in workbook[sheet_name][1]}
                    if any(header not in headers for header in FINAL_RESULT_HEADERS):
                        raise FeishuFileDeliveryError("validate", f"{sheet_name}缺少判断结果或判断来源列")
        finally:
            workbook.close()
        paths[key] = path
    return paths


def _beijing_date(meta: dict[str, Any]) -> str:
    for value in (meta.get("review_completed_at"), meta.get("completed_at"), meta.get("updated_at")):
        try:
            parsed = datetime.fromisoformat(str(value or ""))
        except ValueError:
            continue
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.astimezone(timezone(timedelta(hours=8))).strftime("%Y%m%d")
    task_id = str(meta.get("task_id") or "")
    if re.match(r"^\d{8}", task_id):
        return task_id[:8]
    return datetime.now(timezone(timedelta(hours=8))).strftime("%Y%m%d")


def _baseline(meta: dict[str, Any], version: str) -> str:
    suffix = "40" if version == "4.0" else "51"
    candidates = (
        meta.get(f"baseline_{suffix}"),
        meta.get(f"eea{suffix}_baseline"),
        os.getenv(f"EEA{suffix}_BASELINE", ""),
        meta.get(f"full_compare_{suffix}_parent_url"),
        os.getenv(f"FULL_COMPARE_{suffix}_PARENT_URL", ""),
        OUTPUT_FILENAMES["full_40" if version == "4.0" else "full_51"],
    )
    for candidate in candidates:
        match = _BASELINE_RE.search(str(candidate or ""))
        if match:
            return match.group(1).upper()
    raise FeishuFileDeliveryError("filename", f"无法识别EEA{version}基线，请配置 EEA{suffix}_BASELINE")


def delivery_file_names(meta: dict[str, Any]) -> dict[str, str]:
    date = _beijing_date(meta)
    return {
        "full_40": f"{date}_{_baseline(meta, '4.0')}_EEA4.0全量信号矩阵清单.xlsx",
        "full_51": f"{date}_{_baseline(meta, '5.1')}_EEA5.1全量信号矩阵清单.xlsx",
        "compare_final": f"{date}_EEA4.0和EEA5.1同名信号差异提取.xlsx",
    }


def _stage_files(task_dir: Path, files: dict[str, Path], names: dict[str, str]) -> dict[str, Path]:
    tdir = Path(task_dir).resolve()
    staging_dir = (tdir / "bot" / "feishu_result_files").resolve()
    if staging_dir.parent != (tdir / "bot").resolve():
        raise FeishuFileDeliveryError("stage", "飞书结果暂存路径越界")
    staging_dir.mkdir(parents=True, exist_ok=True)
    staged: dict[str, Path] = {}
    for key in DELIVERY_ORDER:
        target = staging_dir / names[key]
        try:
            shutil.copy2(files[key], target)
        except OSError as exc:
            raise FeishuFileDeliveryError("stage", f"准备飞书结果文件失败：{target.name}：{exc}") from exc
        if not target.is_file() or target.stat().st_size != files[key].stat().st_size:
            raise FeishuFileDeliveryError("stage", f"准备飞书结果文件失败：{target.name}")
        staged[key] = target
    return staged


def _chat_id(meta: dict[str, Any]) -> str:
    chat_id = str(meta.get("feishu_result_chat_id") or meta.get("feishu_chat_id") or os.getenv("FEISHU_RESULT_CHAT_ID", "")).strip()
    if not chat_id.startswith("oc_"):
        raise FeishuFileDeliveryError("target", "缺少有效的 FEISHU_RESULT_CHAT_ID（群chat_id必须以oc_开头）")
    return chat_id


def _client() -> Any:
    mode = os.getenv("FEISHU_FILE_SEND_MODE", "openapi").strip().lower()
    if mode == "lark_cli":
        return LarkCliClient()
    if mode == "openapi":
        return FeishuOpenAPIClient()
    raise FeishuFileDeliveryError("configuration", "FEISHU_FILE_SEND_MODE 只支持 openapi 或 lark_cli")


def _send_file(client: Any, path: Path, chat_id: str) -> dict[str, str]:
    if isinstance(client, LarkCliClient):
        message_id = client.send_file(chat_id=chat_id, file_path=path, timeout=int(os.getenv("FEISHU_FILE_SEND_TIMEOUT_SECONDS", "120")))
        if not message_id:
            raise FeishuFileDeliveryError("send", f"lark-cli发送失败：{path.name}")
        return {"file_name": path.name, "file_key": "", "message_id": str(message_id)}
    result = client.send_file(path, chat_id=chat_id)
    if not isinstance(result, dict) or not result.get("message_id"):
        raise FeishuFileDeliveryError("send", f"飞书文件发送响应无效：{path.name}")
    return {"file_name": path.name, "file_key": str(result.get("file_key") or ""), "message_id": str(result["message_id"])}


def _default_delivery(names: dict[str, str]) -> dict[str, Any]:
    return {
        "status": "pending",
        "attempt_count": 0,
        "started_at": "",
        "completed_at": "",
        "updated_at": "",
        "last_error": "",
        "failed_stage": "",
        "chat_id": "",
        "attachments": {
            key: {
                "status": "pending",
                "display_name": names[key],
                "attempt_count": 0,
                "message_id": "",
                "file_key": "",
                "last_error": "",
            }
            for key in DELIVERY_ORDER
        },
    }


def _normalized_delivery(existing: dict[str, Any], names: dict[str, str]) -> dict[str, Any]:
    delivery = _default_delivery(names)
    delivery.update({key: value for key, value in existing.items() if key != "attachments"})
    for key in DELIVERY_ORDER:
        delivery["attachments"][key].update(dict((existing.get("attachments") or {}).get(key) or {}))
        delivery["attachments"][key]["display_name"] = names[key]
    return delivery


def _save(task_dir: Path, delivery: dict[str, Any]) -> None:
    delivery["updated_at"] = utc_now_iso()
    summary = {"pending": "pending", "sending": "sending", "failed": "failed", "delivered": "delivered"}.get(str(delivery.get("status") or ""), "failed")
    update_task_meta(
        task_dir,
        feishu_file_delivery=delivery,
        result_delivery_status=summary,
        delivery_error=str(delivery.get("last_error") or ""),
    )


def _claim(task_dir: Path) -> Path:
    claim = Path(task_dir) / "bot" / "feishu_file_delivery.lock"
    claim.parent.mkdir(parents=True, exist_ok=True)
    stale_seconds = max(60, int(os.getenv("FEISHU_RESULT_FILE_LOCK_STALE_SECONDS", "600")))
    if claim.exists() and datetime.now().timestamp() - claim.stat().st_mtime > stale_seconds:
        claim.unlink(missing_ok=True)
    try:
        descriptor = os.open(claim, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise FeishuFileDeliveryError("running", "飞书结果文件正在发送") from exc
    with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
        handle.write(f"pid={os.getpid()} at={utc_now_iso()}")
    return claim


def deliver_task_result_files(task_dir: str | Path, *, client: Any | None = None) -> dict[str, Any]:
    """Send exactly three renamed Excel files, resuming only failed items."""

    tdir = Path(task_dir).resolve()
    if not _enabled():
        return {"success": False, "status": "disabled", "last_error": "飞书结果文件发送未启用"}
    try:
        claim = _claim(tdir)
    except FeishuFileDeliveryError as exc:
        return {"success": False, "status": "running", "failed_stage": exc.stage, "last_error": str(exc)}
    try:
        try:
            with get_task_lock(tdir):
                meta = load_task_meta(tdir)
                files = _registered_files(tdir, meta)
                names = delivery_file_names(meta)
                delivery = _normalized_delivery(dict(meta.get("feishu_file_delivery") or {}), names)
                if delivery.get("status") == "delivered" and all(
                    delivery["attachments"][key].get("status") == "success" for key in DELIVERY_ORDER
                ):
                    return {"success": True, **delivery}
                chat_id = _chat_id(meta)
                delivery.update({
                    "status": "sending",
                    "attempt_count": int(delivery.get("attempt_count") or 0) + 1,
                    "started_at": delivery.get("started_at") or utc_now_iso(),
                    "chat_id": chat_id,
                    "last_error": "",
                    "failed_stage": "",
                })
                _save(tdir, delivery)
            staged = _stage_files(tdir, files, names)
            sender = client or _client()
            errors: list[str] = []
            for key in DELIVERY_ORDER:
                attachment = delivery["attachments"][key]
                if attachment.get("status") == "success":
                    continue
                attachment.update({
                    "status": "sending",
                    "attempt_count": int(attachment.get("attempt_count") or 0) + 1,
                    "last_error": "",
                })
                _save(tdir, delivery)
                try:
                    sent = _send_file(sender, staged[key], chat_id)
                except Exception as exc:  # noqa: BLE001 - retain partial progress for retry
                    attachment.update({"status": "failed", "last_error": str(exc)})
                    errors.append(f"{attachment['display_name']}：{exc}")
                    _save(tdir, delivery)
                    continue
                attachment.update({
                    "status": "success",
                    "message_id": sent["message_id"],
                    "file_key": sent.get("file_key", ""),
                    "last_error": "",
                })
                _save(tdir, delivery)

            if errors:
                delivery.update({"status": "failed", "failed_stage": "send", "last_error": "；".join(errors)})
                _save(tdir, delivery)
                return {"success": False, **delivery}
            delivery.update({"status": "delivered", "completed_at": utc_now_iso(), "failed_stage": "", "last_error": ""})
            _save(tdir, delivery)
            update_task_meta(tdir, status="delivered", result_delivered_at=utc_now_iso())
            return {"success": True, **delivery}
        except Exception as exc:  # noqa: BLE001 - persist validation/configuration failures
            stage = exc.stage if isinstance(exc, FeishuFileDeliveryError) else "unexpected"
            meta = load_task_meta(tdir)
            try:
                names = delivery_file_names(meta)
            except FeishuFileDeliveryError:
                date = _beijing_date(meta)
                names = {
                    "full_40": f"{date}_EEA4.0全量信号矩阵清单.xlsx",
                    "full_51": f"{date}_EEA5.1全量信号矩阵清单.xlsx",
                    "compare_final": f"{date}_EEA4.0和EEA5.1同名信号差异提取.xlsx",
                }
            delivery = _normalized_delivery(dict(meta.get("feishu_file_delivery") or {}), names)
            delivery.update({"status": "failed", "failed_stage": stage, "last_error": str(exc)})
            _save(tdir, delivery)
            return {"success": False, **delivery}
    finally:
        claim.unlink(missing_ok=True)
