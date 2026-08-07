"""Export the reviewed result on top of the original two-sheet comparison."""

from __future__ import annotations

import json
from collections import defaultdict, deque
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .ai_review import DIFF_LIST_HEADER, SIGNAL_40_HEADER, SIGNAL_51_HEADER, SOURCE_SHEETS
from .pipeline import OUTPUT_FILENAMES
from .review_store import HISTORY_MANUAL_SOURCE, MANUAL_SOURCE, iter_item_fields


FINAL_REVIEW_FILENAME = "人工审核后最终差异结果.xlsx"
RESULT_HEADER = "判断结果"
SOURCE_HEADER = "判断来源"
FINAL_RESULT_HEADERS = (RESULT_HEADER, SOURCE_HEADER)
SUMMARY_HEADERS = (
    "4.0_发送ECU汇总",
    "5.1_发送ECU汇总",
    "4.0_接收ECU汇总",
    "5.1_接收ECU汇总",
)
SOURCE_FILE_HEADERS = ("4.0_信号来源文件", "5.1_信号来源文件")
DESCRIPTION_ANCHOR_HEADER = "5.1_信号值描述"
FINAL_COLUMN_WIDTHS = {
    SIGNAL_40_HEADER: 24,
    SIGNAL_51_HEADER: 24,
    "去前缀后匹配名": 24,
    DIFF_LIST_HEADER: 48,
    RESULT_HEADER: 12,
    SOURCE_HEADER: 12,
    "4.0_单位": 12,
    "5.1_单位": 12,
    "4.0_信号值描述": 36,
    "5.1_信号值描述": 36,
    "4.0_发送ECU汇总": 24,
    "5.1_发送ECU汇总": 24,
    "4.0_接收ECU汇总": 24,
    "5.1_接收ECU汇总": 24,
    "4.0_ECU收发状态_原始": 20,
    "5.1_ECU收发状态_原始": 20,
    "4.0_ECU收发状态_标准化": 18,
    "5.1_ECU收发状态_标准化": 18,
    "4.0_信号来源文件": 36,
    "5.1_信号来源文件": 36,
}
# Backward-compatible name used by older diagnostics; final delivery now keeps
# exactly the original two comparison sheets.
SHEET_RULES = list(SOURCE_SHEETS)


def _load_json(path: Path) -> Any:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def _item_decision(item: dict[str, Any], state_items: dict[str, Any]) -> tuple[str, str]:
    item_id = str(item.get("item_id") or "")
    reviews = dict((state_items.get(item_id) or {}).get("field_reviews") or {})
    results: list[str] = []
    sources: list[str] = []
    for diff in iter_item_fields(item):
        field_key = str(diff.get("field_key") or "")
        review = dict(reviews.get(field_key) or {})
        result = str(review.get("result") or "")
        if not review.get("reviewed") or result not in {"same", "different"}:
            raise ValueError(f"审核项尚未完成：{item_id}::{field_key}")
        results.append(result)
        sources.append(str(review.get("decision_source") or ""))
    if not results:
        raise ValueError(f"审核项缺少差异字段：{item_id}")
    final_result = "不同" if "different" in results else "相同"
    final_source = "人工" if any(source in {MANUAL_SOURCE, HISTORY_MANUAL_SOURCE} for source in sources) else "系统"
    return final_result, final_source


def _decision_queues(items: list[dict[str, Any]], state_items: dict[str, Any]) -> dict[tuple[str, str, str], deque[tuple[str, str]]]:
    decisions: dict[tuple[str, str, str], deque[tuple[str, str]]] = defaultdict(deque)
    for item in items:
        key = (
            str(item.get("source_sheet") or ""),
            str(item.get("signal_40") or ""),
            str(item.get("signal_51") or ""),
        )
        decisions[key].append(_item_decision(item, state_items))
    return decisions


def _header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): index
        for index, cell in enumerate(ws[1], start=1)
        if cell.value is not None
    }


def _ensure_result_columns(ws) -> tuple[int, int]:
    headers = _header_map(ws)
    missing = [name for name in (DIFF_LIST_HEADER, SIGNAL_40_HEADER, SIGNAL_51_HEADER) if name not in headers]
    if missing:
        raise ValueError(f"Sheet {ws.title} 缺少列：{'、'.join(missing)}")
    result_col = headers.get(RESULT_HEADER)
    source_col = headers.get(SOURCE_HEADER)
    if result_col and source_col:
        return result_col, source_col
    if result_col or source_col:
        raise ValueError(f"Sheet {ws.title} 的判断列结构不完整")
    insert_at = headers[DIFF_LIST_HEADER] + 1
    ws.insert_cols(insert_at, amount=2)
    ws.cell(1, insert_at, RESULT_HEADER)
    ws.cell(1, insert_at + 1, SOURCE_HEADER)
    return insert_at, insert_at + 1


def _style_result_columns(ws, result_col: int, source_col: int) -> None:
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_styles = {
        result_col: PatternFill("solid", fgColor="C65911"),
        source_col: PatternFill("solid", fgColor="548235"),
    }
    body_styles = {
        result_col: PatternFill("solid", fgColor="FCE4D6"),
        source_col: PatternFill("solid", fgColor="E2F0D9"),
    }
    for column in (result_col, source_col):
        header = ws.cell(1, column)
        header.fill = header_styles[column]
        header.font = Font(bold=True, color="FFFFFF")
        header.border = border
        header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in range(2, (ws.max_row or 1) + 1):
            cell = ws.cell(row, column)
            cell.fill = body_styles[column]
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[ws.cell(1, result_col).column_letter].width = 12
    ws.column_dimensions[ws.cell(1, source_col).column_letter].width = 12
    ws.auto_filter.ref = ws.dimensions


def _compact_final_columns(ws) -> None:
    headers = _header_map(ws)
    for header, width in FINAL_COLUMN_WIDTHS.items():
        column = headers.get(header)
        if column:
            ws.column_dimensions[ws.cell(1, column).column_letter].width = width


def _reorder_final_columns(ws) -> None:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    if len(headers) != len(set(headers)):
        raise ValueError(f"Sheet {ws.title} 存在重复列名，无法重排")
    required = (DESCRIPTION_ANCHOR_HEADER, *SUMMARY_HEADERS, *SOURCE_FILE_HEADERS)
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"Sheet {ws.title} 缺少最终结果重排列：{'、'.join(missing)}")

    moved = set(SUMMARY_HEADERS + SOURCE_FILE_HEADERS)
    desired = [header for header in headers if header not in moved]
    insert_at = desired.index(DESCRIPTION_ANCHOR_HEADER) + 1
    desired[insert_at:insert_at] = list(SUMMARY_HEADERS)
    desired.extend(SOURCE_FILE_HEADERS)
    if desired == headers:
        return

    source_indexes = {header: index for index, header in enumerate(headers, start=1)}
    cells: dict[str, list[tuple[Any, Any, Any, Any]]] = {}
    dimensions: dict[str, dict[str, Any]] = {}
    for header, source_index in source_indexes.items():
        cells[header] = [
            (cell.value, copy(cell._style), copy(cell.hyperlink), copy(cell.comment))
            for cell in (ws.cell(row, source_index) for row in range(1, (ws.max_row or 1) + 1))
        ]
        source_dimension = ws.column_dimensions[ws.cell(1, source_index).column_letter]
        dimensions[header] = {
            "width": source_dimension.width,
            "hidden": source_dimension.hidden,
            "bestFit": source_dimension.bestFit,
            "outlineLevel": source_dimension.outlineLevel,
            "collapsed": source_dimension.collapsed,
        }

    for destination_index, header in enumerate(desired, start=1):
        for row, (value, style, hyperlink, comment) in enumerate(cells[header], start=1):
            target = ws.cell(row, destination_index)
            target.value = value
            target._style = copy(style)
            target.hyperlink = copy(hyperlink)
            target.comment = copy(comment)
        destination_dimension = ws.column_dimensions[ws.cell(1, destination_index).column_letter]
        for attribute, value in dimensions[header].items():
            setattr(destination_dimension, attribute, value)


def _source_compare_path(output_path: Path, compare_file_path: Path | None) -> Path:
    source = Path(compare_file_path) if compare_file_path is not None else output_path.parent / OUTPUT_FILENAMES["compare"]
    if not source.is_file():
        raise FileNotFoundError(f"原始同名信号差异文件不存在：{source}")
    if source.resolve() == output_path.resolve():
        raise ValueError("最终结果文件不能覆盖原始同名信号差异文件")
    return source


def export_final_review_result(
    review_items_path: Path,
    review_state_path: Path,
    output_file_path: Path,
    *,
    compare_file_path: Path | None = None,
) -> dict[str, int]:
    """Copy the source comparison, remove the AI sheet and append final decisions."""

    items = _load_json(review_items_path)
    state = _load_json(review_state_path)
    decisions = _decision_queues(items, dict(state.get("items") or {}))
    output_path = Path(output_file_path)
    source_path = _source_compare_path(output_path, compare_file_path)
    workbook = load_workbook(source_path)
    try:
        missing_sheets = [name for name in SOURCE_SHEETS if name not in workbook.sheetnames]
        if missing_sheets:
            raise ValueError(f"原始同名信号差异文件缺少Sheet：{'、'.join(missing_sheets)}")
        for sheet_name in list(workbook.sheetnames):
            if sheet_name not in SOURCE_SHEETS:
                del workbook[sheet_name]

        stats = {"判断结果-相同": 0, "判断结果-不同": 0, "判断来源-人工": 0, "判断来源-系统": 0, "审核信号总数": 0}
        for sheet_name in SOURCE_SHEETS:
            ws = workbook[sheet_name]
            result_col, source_col = _ensure_result_columns(ws)
            headers = _header_map(ws)
            same_rows: list[int] = []
            for row in range(2, (ws.max_row or 1) + 1):
                key = (
                    sheet_name,
                    str(ws.cell(row, headers[SIGNAL_40_HEADER]).value or "").strip(),
                    str(ws.cell(row, headers[SIGNAL_51_HEADER]).value or "").strip(),
                )
                queue = decisions.get(key)
                if not queue:
                    raise ValueError(f"原始差异行未找到审核结论：{sheet_name} 第{row}行 {key[1]} / {key[2]}")
                result, source = queue.popleft()
                ws.cell(row, result_col, result)
                ws.cell(row, source_col, source)
                stats[f"判断结果-{result}"] += 1
                stats[f"判断来源-{source}"] += 1
                stats["审核信号总数"] += 1
                if result == "相同":
                    same_rows.append(row)
            for row in reversed(same_rows):
                ws.delete_rows(row)
            _style_result_columns(ws, result_col, source_col)
            _reorder_final_columns(ws)
            _compact_final_columns(ws)

        remaining = sum(len(queue) for queue in decisions.values())
        if remaining:
            raise ValueError(f"有 {remaining} 条审核结论未能匹配原始差异行")
        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return stats
    finally:
        workbook.close()
