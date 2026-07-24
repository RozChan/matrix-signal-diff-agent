"""Persistent storage helpers for the local signal-level human review workflow."""

from __future__ import annotations

import hashlib
import json
import os
import tempfile
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from .task_lock import get_task_lock

AI_REVIEW_SHEET = "AI辅助复核与人工审核明细"
REVIEW_ITEMS_FILE = "review_items.json"
REVIEW_STATE_FILE = "review_state.json"
REVIEW_LOG_FILE = "review_log.jsonl"
TASK_META_FILE = "task_meta.json"

FIELD_REVIEW_RESULTS = {"same", "different"}
MANUAL_TEXT_FIELDS = {"信号值描述", "单位"}
SIGNAL_AI_JUDGEMENTS = ["真实差异", "疑似可忽略", "无法判断", "未启用"]
SYSTEM_DEFAULT_SOURCE = "system_default"
MANUAL_SOURCE = "manual"
HISTORY_MANUAL_SOURCE = "history_manual"

ITEM_HEADER_MAP = {
    "来源Sheet": "source_sheet",
    "4.0信号名": "signal_40",
    "5.1信号名": "signal_51",
    "差异字段汇总": "diff_fields_text",
    "差异字段数量": "diff_field_count",
    "是否包含数值类差异": "has_numeric_diff",
    "是否包含文本类差异": "has_text_diff",
    "原始差异点list": "original_diff_list",
    "字段差异明细": "field_diff_details",
    "信号级AI判断结果": "signal_ai_judgement",
    "差异类型汇总": "difference_type_summary",
    "置信度": "confidence",
    "信号级AI判断理由": "signal_ai_reason",
    "信号级AI建议处理方式": "signal_ai_suggested_action",
}


def utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


class ReviewConflictError(RuntimeError):
    pass


class ReviewLockError(RuntimeError):
    pass


def _parse_time(value: Any) -> datetime | None:
    try:
        if not value:
            return None
        parsed = datetime.fromisoformat(str(value))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)
    except ValueError:
        return None


def _review_lock_minutes() -> int:
    import os

    return int(os.getenv("REVIEW_LOCK_TIMEOUT_MINUTES", "30"))


def _task_dir_from_review_dir(review_dir: Path) -> Path:
    path = Path(review_dir)
    return path.parent if path.name == "review" else path


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def _atomic_write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(data, ensure_ascii=False, indent=2)
    last_error: PermissionError | None = None
    for attempt in range(5):
        tmp_name = ""
        fd = -1
        try:
            fd, tmp_name = tempfile.mkstemp(prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
            with os.fdopen(fd, "w", encoding="utf-8") as fh:
                fd = -1
                fh.write(payload)
                fh.flush()
                os.fsync(fh.fileno())
            os.replace(tmp_name, path)
            return
        except PermissionError as exc:
            last_error = exc
            if fd >= 0:
                os.close(fd)
            if tmp_name:
                Path(tmp_name).unlink(missing_ok=True)
            if attempt == 4:
                raise
            time.sleep(0.05 * (2**attempt))
        except Exception:
            if fd >= 0:
                os.close(fd)
            if tmp_name:
                Path(tmp_name).unlink(missing_ok=True)
            raise
    if last_error:
        raise last_error


def _task_meta_path(task_dir: Path) -> Path:
    return Path(task_dir) / TASK_META_FILE


def _review_items_path(review_dir: Path) -> Path:
    return Path(review_dir) / REVIEW_ITEMS_FILE


def _review_state_path(review_dir: Path) -> Path:
    return Path(review_dir) / REVIEW_STATE_FILE


def _review_log_path(review_dir: Path) -> Path:
    return Path(review_dir) / REVIEW_LOG_FILE


def create_task_meta(task_dir: Path, task_id: str, input_40_count: int = 0, input_51_count: int = 0, status: str = "created") -> dict[str, Any]:
    with get_task_lock(Path(task_dir)):
        task_path = Path(task_dir)
        output_dir = task_path / "output"
        review_dir = task_path / "review"
        now = utc_now_iso()
        meta = {
            "task_id": task_id,
            "created_at": now,
            "updated_at": now,
            "status": status,
            "input_40_count": input_40_count,
            "input_51_count": input_51_count,
            "output_dir": str(output_dir),
            "review_items_path": str(review_dir / REVIEW_ITEMS_FILE),
            "review_state_path": str(review_dir / REVIEW_STATE_FILE),
            "final_review_file": str(output_dir / "人工审核后最终差异结果.xlsx"),
            "error": "",
            "review_lock_status": "unlocked",
            "review_owner": "",
            "review_session_id": "",
            "review_locked_at": "",
            "review_lock_last_active_at": "",
            "review_lock_expires_at": "",
            "review_completed": False,
            "review_completed_at": "",
            "final_generation_status": "",
        }
        save_task_meta(task_path, meta)
        return meta


def load_task_meta(task_dir: Path) -> dict[str, Any]:
    return _read_json(_task_meta_path(Path(task_dir)), {})


def save_task_meta(task_dir: Path, meta: dict[str, Any]) -> None:
    with get_task_lock(Path(task_dir)):
        meta = dict(meta)
        meta["updated_at"] = utc_now_iso()
        _atomic_write_json(_task_meta_path(Path(task_dir)), meta)


def update_task_meta(task_dir: Path, **updates: Any) -> dict[str, Any]:
    with get_task_lock(Path(task_dir)):
        meta = load_task_meta(task_dir)
        meta.update(updates)
        save_task_meta(task_dir, meta)
        return meta


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip() in {"是", "true", "True", "1"}


def _split_fields(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    return [part.strip() for part in text.replace(",", "、").split("、") if part.strip()]


def _field_type_from_detail(line: str) -> str:
    if "数值类差异" in line:
        return "numeric"
    if "文本类差异" in line:
        return "text"
    return "unknown"


def _parse_field_diff_details(text: str, diff_fields: list[str]) -> list[dict[str, str]]:
    details = []
    if text:
        blocks = [block.strip() for block in text.split("\n\n") if block.strip()]
        for block in blocks:
            lines = block.splitlines()
            field = lines[0].strip("【】") if lines else "未解析"
            value_40_lines: list[str] = []
            value_51_lines: list[str] = []
            field_type = "unknown"
            current_target = ""
            for line in lines[1:]:
                if line.startswith("4.0："):
                    current_target = "4.0"
                    value_40_lines.append(line.replace("4.0：", "", 1))
                elif line.startswith("5.1："):
                    current_target = "5.1"
                    value_51_lines.append(line.replace("5.1：", "", 1))
                elif line.startswith("类型："):
                    current_target = ""
                    field_type = _field_type_from_detail(line)
                elif current_target == "4.0":
                    value_40_lines.append(line)
                elif current_target == "5.1":
                    value_51_lines.append(line)
            details.append({"diff_field": field, "value_40": "\n".join(value_40_lines), "value_51": "\n".join(value_51_lines), "field_type": field_type})
    if details:
        return details
    return [{"diff_field": field, "value_40": "", "value_51": "", "field_type": "unknown"} for field in diff_fields]


def build_item_id(item: dict[str, Any]) -> str:
    parts = [item.get("source_sheet", ""), item.get("signal_40", ""), item.get("signal_51", ""), item.get("original_diff_list", "")]
    raw = "\u241f".join(str(part) for part in parts)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def is_signal_level_item(item: dict[str, Any]) -> bool:
    return "field_diffs" in item and "signal_ai_judgement" in item


def _header_map(ws) -> dict[str, int]:
    return {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], start=1) if cell.value is not None}


def _cell(ws, row_idx: int, headers: dict[str, int], title: str) -> str:
    col = headers.get(title)
    if not col:
        return ""
    value = ws.cell(row=row_idx, column=col).value
    return "" if value is None else str(value).strip()


def generate_review_items_from_excel(compare_file_path: Path, review_dir: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    review_path = Path(review_dir)
    review_path.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(compare_file_path, read_only=True, data_only=True)
    try:
        if AI_REVIEW_SHEET not in wb.sheetnames:
            raise ValueError(f"最终差异文件缺少 sheet：{AI_REVIEW_SHEET}")
        ws = wb[AI_REVIEW_SHEET]
        headers = _header_map(ws)
        items: list[dict[str, Any]] = []
        seen: dict[str, int] = {}
        for row_idx in range(2, (ws.max_row or 1) + 1):
            raw = {field: _cell(ws, row_idx, headers, title) for title, field in ITEM_HEADER_MAP.items()}
            if not any(raw.values()):
                continue
            diff_fields = _split_fields(raw.get("diff_fields_text"))
            field_diffs = _parse_field_diff_details(raw.get("field_diff_details", ""), diff_fields)
            item: dict[str, Any] = {
                "source_sheet": raw.get("source_sheet", ""),
                "signal_40": raw.get("signal_40", ""),
                "signal_51": raw.get("signal_51", ""),
                "diff_fields": diff_fields,
                "diff_field_count": int(raw.get("diff_field_count") or len(field_diffs)),
                "has_numeric_diff": _as_bool(raw.get("has_numeric_diff")),
                "has_text_diff": _as_bool(raw.get("has_text_diff")),
                "original_diff_list": raw.get("original_diff_list", ""),
                "field_diffs": field_diffs,
                "field_diff_details": raw.get("field_diff_details", ""),
                "signal_ai_judgement": raw.get("signal_ai_judgement", ""),
                "difference_type_summary": raw.get("difference_type_summary", ""),
                "confidence": raw.get("confidence", ""),
                "signal_ai_reason": raw.get("signal_ai_reason", ""),
                "signal_ai_suggested_action": raw.get("signal_ai_suggested_action", ""),
            }
            base_id = build_item_id(item)
            count = seen.get(base_id, 0)
            seen[base_id] = count + 1
            item["item_id"] = base_id if count == 0 else f"{base_id}_{count + 1}"
            items.append(item)
    finally:
        wb.close()

    _atomic_write_json(_review_items_path(review_path), items)
    return items


def load_review_items(review_dir: Path) -> list[dict[str, Any]]:
    return _read_json(_review_items_path(Path(review_dir)), [])


def iter_item_fields(item: dict[str, Any]) -> list[dict[str, Any]]:
    """Return stable field identities for field-level review and export."""

    counts: dict[str, int] = {}
    fields: list[dict[str, Any]] = []
    for position, raw in enumerate(item.get("field_diffs") or [], start=1):
        diff = dict(raw)
        name = str(diff.get("diff_field") or "未解析")
        counts[name] = counts.get(name, 0) + 1
        key = name if counts[name] == 1 else f"{name}#{counts[name]}"
        diff.update(field_key=key, position=position)
        fields.append(diff)
    return fields


def is_manual_text_review_item(item: dict[str, Any]) -> bool:
    """Return true only for description/unit-only signals without other differences."""

    fields = iter_item_fields(item)
    return bool(fields) and all(
        diff.get("field_type") == "text" and diff.get("diff_field") in MANUAL_TEXT_FIELDS
        for diff in fields
    )


def manual_text_review_kind(item: dict[str, Any]) -> str:
    if not is_manual_text_review_item(item):
        return ""
    names = {str(diff.get("diff_field") or "") for diff in iter_item_fields(item)}
    if names == {"信号值描述"}:
        return "description_only"
    if names == {"单位"}:
        return "unit_only"
    if names == MANUAL_TEXT_FIELDS:
        return "description_and_unit"
    return ""


def get_default_review_state(item: dict[str, Any]) -> dict[str, Any]:
    now = utc_now_iso()
    system_different = any(diff.get("field_type") == "numeric" for diff in iter_item_fields(item))
    field_reviews: dict[str, dict[str, Any]] = {}
    for diff in iter_item_fields(item):
        field_reviews[diff["field_key"]] = {
            "diff_field": diff.get("diff_field") or "未解析",
            "result": "different" if system_different else "",
            "reviewed": system_different,
            "decision_source": SYSTEM_DEFAULT_SOURCE if system_different else "",
            "reviewed_at": now if system_different else "",
            "updated_at": now,
            "reviewer": "",
        }
    return {"field_reviews": field_reviews, "updated_at": now}


def _normalize_review_entry(entry: dict[str, Any], item: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized = dict(entry) if isinstance(entry, dict) else {}
    default = get_default_review_state(item or {})
    normalized.setdefault("updated_at", normalized.get("reviewed_at") or utc_now_iso())
    existing_fields = normalized.get("field_reviews") if isinstance(normalized.get("field_reviews"), dict) else {}
    merged = dict(default.get("field_reviews", {}))
    for key, value in existing_fields.items():
        if isinstance(value, dict):
            current = dict(merged.get(key, {}))
            if current.get("decision_source") != SYSTEM_DEFAULT_SOURCE:
                current.update(value)
            current["result"] = str(current.get("result") or "")
            current["reviewed"] = bool(current.get("reviewed") and current["result"] in FIELD_REVIEW_RESULTS)
            merged[key] = current
    normalized["field_reviews"] = merged
    return normalized


def init_review_state(review_dir: Path, task_id: str, items: list[dict[str, Any]] | None = None, overwrite: bool = False) -> dict[str, Any]:
    review_path = Path(review_dir)
    item_list = items if items is not None else load_review_items(review_path)
    existing = load_review_state(review_path) if _review_state_path(review_path).exists() and not overwrite else {"items": {}}
    state = {"schema_version": 2, "task_id": existing.get("task_id") or task_id, "updated_at": existing.get("updated_at") or utc_now_iso(), "revision": int(existing.get("revision") or 0), "items": dict(existing.get("items", {}))}
    changed = overwrite or not _review_state_path(review_path).exists()
    for item in item_list:
        item_id = item.get("item_id")
        if not item_id:
            continue
        current = state["items"].get(item_id)
        if current is None or overwrite:
            state["items"][item_id] = get_default_review_state(item)
            changed = True
        else:
            normalized = _normalize_review_entry(current, item)
            if normalized != current:
                state["items"][item_id] = normalized
                changed = True
    if changed:
        save_review_state(review_path, state)
    return apply_review_history(review_path, task_id, item_list, load_review_state(review_path))


def apply_review_history(review_dir: Path, task_id: str, items: list[dict[str, Any]], state: dict[str, Any]) -> dict[str, Any]:
    """Fill still-pending description/unit fields from exact cross-task history."""

    from .review_history import history_database_path, lookup_history_decision

    reused: list[dict[str, str]] = []
    state_items = state.setdefault("items", {})
    db_path = history_database_path(review_dir)
    for item in items:
        if not is_manual_text_review_item(item):
            continue
        item_id = str(item.get("item_id") or "")
        item_review = state_items.get(item_id, {})
        reviews = item_review.get("field_reviews", {})
        for diff in iter_item_fields(item):
            field_key = str(diff.get("field_key") or "")
            field_review = reviews.get(field_key, {})
            if field_review.get("reviewed"):
                continue
            historical = lookup_history_decision(item, diff, db_path=db_path)
            if not historical:
                continue
            now = utc_now_iso()
            reviews[field_key] = {
                **field_review,
                "result": historical["result"],
                "reviewed": True,
                "decision_source": HISTORY_MANUAL_SOURCE,
                "reviewed_at": now,
                "updated_at": now,
                "reviewer": historical.get("latest_reviewer", ""),
                "history_fingerprint": historical["fingerprint"],
                "history_task_id": historical.get("latest_task_id", ""),
                "history_confirmed_at": historical.get("latest_confirmed_at", ""),
            }
            reused.append({"item_id": item_id, "field_key": field_key, "result": historical["result"]})
        item_review["field_reviews"] = reviews
        state_items[item_id] = item_review
    if not reused:
        return state
    state["history_reused_count"] = int(state.get("history_reused_count") or 0) + len(reused)
    save_review_state(review_dir, state, increment_revision=True)
    for entry in reused:
        try:
            append_review_log(review_dir, {"task_id": task_id, "action": "apply_review_history", **entry})
        except OSError:
            pass
    return load_review_state(review_dir)


def load_review_state(review_dir: Path) -> dict[str, Any]:
    state = _read_json(_review_state_path(Path(review_dir)), {"task_id": "", "updated_at": "", "items": {}})
    state.setdefault("task_id", "")
    state.setdefault("updated_at", "")
    state.setdefault("revision", 0)
    state.setdefault("schema_version", 2)
    state.setdefault("items", {})
    state["items"] = {item_id: _normalize_review_entry(entry) for item_id, entry in state.get("items", {}).items()}
    return state


def save_review_state(review_dir: Path, state: dict[str, Any], *, increment_revision: bool = False) -> None:
    with get_task_lock(_task_dir_from_review_dir(Path(review_dir))):
        state = dict(state)
        state["updated_at"] = utc_now_iso()
        state["revision"] = int(state.get("revision") or 0) + (1 if increment_revision else 0)
        state.setdefault("items", {})
        _atomic_write_json(_review_state_path(Path(review_dir)), state)


def append_review_log(review_dir: Path, entry: dict[str, Any]) -> None:
    log_entry = {"time": utc_now_iso(), **entry}
    log_path = _review_log_path(Path(review_dir))
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(log_entry, ensure_ascii=False) + "\n")


def update_review_field(
    review_dir: Path,
    task_id: str,
    item_id: str,
    field_key: str,
    result: str,
    reviewer: str = "",
    *,
    base_revision: int | None = None,
    session_id: str = "",
) -> dict[str, Any]:
    if result not in FIELD_REVIEW_RESULTS:
        raise ValueError(f"不支持的字段确认结果：{result}")
    with get_task_lock(_task_dir_from_review_dir(Path(review_dir))):
        if session_id:
            ensure_review_lock_holder(_task_dir_from_review_dir(Path(review_dir)), session_id)
        state = load_review_state(review_dir)
        if base_revision is not None and int(state.get("revision") or 0) != int(base_revision):
            raise ReviewConflictError("审核数据已被其他用户更新，请刷新页面")
        if not state.get("task_id"):
            state["task_id"] = task_id
        item_review = _normalize_review_entry(state.setdefault("items", {}).get(item_id, {}))
        if field_key not in item_review.get("field_reviews", {}):
            raise ValueError(f"审核字段不存在：{field_key}")
        previous = dict(item_review["field_reviews"][field_key])
        now = utc_now_iso()
        item_review["field_reviews"][field_key] = {
            **previous,
            "result": result,
            "reviewed": True,
            "decision_source": MANUAL_SOURCE,
            "reviewed_at": now,
            "updated_at": now,
            "reviewer": reviewer,
        }
        item_review["updated_at"] = now
        state["items"][item_id] = item_review
        save_review_state(review_dir, state, increment_revision=True)
        state = load_review_state(review_dir)
    try:
        append_review_log(review_dir, {"task_id": task_id, "item_id": item_id, "field_key": field_key, "diff_field": previous.get("diff_field", ""), "action": "update_field_review", "old_result": previous.get("result", ""), "new_result": result, "reviewer": reviewer})
    except OSError:
        pass
    return state


def acquire_review_lock(task_dir: Path, session_id: str, owner: str = "", *, takeover: bool = False) -> dict[str, Any]:
    with get_task_lock(Path(task_dir)):
        meta = load_task_meta(task_dir)
        if meta.get("review_completed") or meta.get("status") in {"final_exported", "delivered"}:
            raise ReviewLockError("该任务已完成审核，不能继续编辑")
        now = datetime.now(timezone.utc)
        now_iso = utc_now_iso()
        expires = _parse_time(meta.get("review_lock_expires_at"))
        current_session = str(meta.get("review_session_id") or "")
        locked = meta.get("review_lock_status") == "locked" and expires and expires > now
        if locked and current_session != session_id and not takeover:
            raise ReviewLockError(f"该任务正在由{meta.get('review_owner') or current_session}审核，当前为只读模式")
        is_takeover = bool(takeover and current_session and current_session != session_id)
        lock_started_at = meta.get("review_locked_at") if current_session == session_id and locked else now_iso
        updates = {
            "review_lock_status": "locked",
            "review_owner": owner or session_id,
            "review_session_id": session_id,
            "review_locked_at": lock_started_at,
            "review_lock_last_active_at": now_iso,
            "review_lock_expires_at": (now + timedelta(minutes=_review_lock_minutes())).isoformat(timespec="seconds"),
        }
        if is_takeover:
            updates.update(
                review_takeover_at=now_iso,
                review_takeover_from_session=current_session,
                review_takeover_to_session=session_id,
            )
        return update_task_meta(task_dir, **updates)


def heartbeat_review_lock(task_dir: Path, session_id: str) -> dict[str, Any]:
    with get_task_lock(Path(task_dir)):
        ensure_review_lock_holder(task_dir, session_id)
        return update_task_meta(
            task_dir,
            review_lock_last_active_at=utc_now_iso(),
            review_lock_expires_at=(datetime.now(timezone.utc) + timedelta(minutes=_review_lock_minutes())).isoformat(timespec="seconds"),
        )


def ensure_review_lock_holder(task_dir: Path, session_id: str) -> None:
    meta = load_task_meta(task_dir)
    if meta.get("review_completed") or meta.get("status") in {"final_exported", "delivered"}:
        raise ReviewLockError("该任务已完成审核或已交付，不能继续修改")
    if meta.get("review_lock_status") != "locked" or meta.get("review_session_id") != session_id:
        raise ReviewLockError("当前会话未持有审核编辑锁")
    expires = _parse_time(meta.get("review_lock_expires_at"))
    if expires and expires <= datetime.now(timezone.utc):
        raise ReviewLockError("审核编辑锁已过期，请重新获取")


def begin_final_generation(task_dir: Path, session_id: str) -> dict[str, Any]:
    with get_task_lock(Path(task_dir)):
        ensure_review_lock_holder(task_dir, session_id)
        meta = load_task_meta(task_dir)
        if meta.get("review_completed") or meta.get("final_generation_status") == "generating" or meta.get("status") in {"final_exported", "delivered"} or meta.get("result_delivery_status") in {"sending", "delivered"}:
            raise ReviewLockError("该任务已完成审核或正在生成结果")
        return update_task_meta(task_dir, review_completed=True, review_completed_at=utc_now_iso(), final_generation_status="generating")


def review_badge(item: dict[str, Any], review: dict[str, Any]) -> str:
    fields = review.get("field_reviews", {})
    if any(value.get("decision_source") in {MANUAL_SOURCE, HISTORY_MANUAL_SOURCE} for value in fields.values()):
        return "人工已确认"
    if all(value.get("reviewed") for value in fields.values()):
        return "系统判定"
    return "待人工确认"


def review_sort_key(item: dict[str, Any], review: dict[str, Any]) -> tuple[int, str, str, str]:
    fields = review.get("field_reviews", {})
    pending = any(not value.get("reviewed") for value in fields.values())
    priority = 0 if pending else 1
    return (priority, item.get("source_sheet", ""), item.get("signal_40", ""), item.get("signal_51", ""))


def compute_review_stats(items: list[dict[str, Any]], state: dict[str, Any]) -> dict[str, int | float | str]:
    state_items = state.get("items", {}) if isinstance(state, dict) else {}
    total_fields = sum(len(iter_item_fields(item)) for item in items)
    stats: dict[str, int | float | str] = {
        "signal_total": len(items), "field_total": total_fields, "pending_manual": 0,
        "manual_same": 0, "manual_different": 0, "system_different": 0,
        "manual_confirmed": 0, "history_reused": 0, "description_only_signals": 0, "unit_only_signals": 0,
        "description_and_unit_signals": 0, "numeric_difference_signals": 0,
        "updated_at": state.get("updated_at", "") if isinstance(state, dict) else "",
    }
    for item in items:
        kind = manual_text_review_kind(item)
        if kind:
            stats[f"{kind}_signals"] += 1
        if any(diff.get("field_type") == "numeric" for diff in iter_item_fields(item)):
            stats["numeric_difference_signals"] += 1
        review = _normalize_review_entry(state_items.get(item.get("item_id"), {}), item)
        for field_review in review.get("field_reviews", {}).values():
            result = field_review.get("result", "")
            source = field_review.get("decision_source", "")
            if not field_review.get("reviewed"):
                stats["pending_manual"] += 1
            elif source == SYSTEM_DEFAULT_SOURCE and result == "different":
                stats["system_different"] += 1
            elif source in {MANUAL_SOURCE, HISTORY_MANUAL_SOURCE}:
                stats["manual_confirmed"] += 1
                stats["manual_same" if result == "same" else "manual_different"] += 1
                if source == HISTORY_MANUAL_SOURCE:
                    stats["history_reused"] += 1
    return stats
