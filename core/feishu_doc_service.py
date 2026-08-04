"""Publish the three registered final workbooks to a Feishu cloud document."""

from __future__ import annotations

import json
import html
import os
import re
import shutil
import subprocess
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .final_export import FINAL_REVIEW_FILENAME, SHEET_RULES
from .pipeline import OUTPUT_FILENAMES
from .review_store import compute_review_stats, load_review_items, load_review_state, load_task_meta, update_task_meta, utc_now_iso
from .task_lock import get_task_lock

ATTACHMENT_ORDER = ("full_40", "full_51", "compare_final")
ATTACHMENT_LABELS = {
    "full_40": "EEA4.0全量信号矩阵清单.xlsx",
    "full_51": "EEA5.1全量信号矩阵清单.xlsx",
    "compare_final": "人工审核后最终差异结果.xlsx",
}
DELIVERY_STAGING_DIR = "feishu_delivery_attachments"


class FeishuDocumentError(RuntimeError):
    def __init__(self, error_type: str, message: str, *, stdout: str = "", stderr: str = "") -> None:
        self.error_type = error_type
        self.stdout = stdout
        self.stderr = stderr
        super().__init__(message)


def _enabled() -> bool:
    return os.getenv("FEISHU_DOC_DELIVERY_ENABLED", "false").strip().lower() == "true"


def _redact(text: str) -> str:
    cleaned = str(text or "")
    for name in ("FEISHU_CUSTOM_BOT_WEBHOOK", "FEISHU_CUSTOM_BOT_SECRET", "FEISHU_RESULT_FOLDER_TOKEN"):
        value = os.getenv(name, "").strip()
        if value:
            cleaned = cleaned.replace(value, f"<{name.lower()}-redacted>")
    return re.sub(r"(?i)(tenant_access_token|user_access_token|refresh_token)\s*[:=]\s*[^\s,}]+", r"\1=<redacted>", cleaned)


def extract_last_json_object(output: str) -> dict[str, Any]:
    """Extract the last complete JSON object from mixed lark-cli output."""

    decoder = json.JSONDecoder()
    objects: list[tuple[int, int, dict[str, Any]]] = []
    for index, char in enumerate(output or ""):
        if char != "{":
            continue
        try:
            value, end = decoder.raw_decode(output[index:])
        except json.JSONDecodeError:
            continue
        if isinstance(value, dict):
            objects.append((index, index + end, value))
    if not objects:
        raise FeishuDocumentError("cli_output_parse_failed", "无法从lark-cli输出中解析JSON", stdout=_redact(output))
    return max(objects, key=lambda entry: (entry[1], -entry[0]))[2]


def check_lark_cli_environment() -> dict[str, Any]:
    cli_path = Path(os.getenv("LARK_CLI_PATH", "").strip())
    folder_token = os.getenv("FEISHU_RESULT_FOLDER_TOKEN", "").strip()
    if not str(cli_path):
        raise FeishuDocumentError("lark_cli_not_found", "LARK_CLI_PATH未配置")
    if not cli_path.is_file():
        raise FeishuDocumentError("lark_cli_not_found", f"lark-cli不存在：{cli_path}")
    if not folder_token:
        raise FeishuDocumentError("folder_permission_denied", "FEISHU_RESULT_FOLDER_TOKEN未配置")
    return {"success": True, "cli_path": str(cli_path), "parent_token": folder_token}


def _classify_cli_error(stderr: str, default: str) -> str:
    text = str(stderr or "").lower()
    if "need_user_authorization" in text or "token_missing" in text or "authorization" in text:
        return "user_authorization_missing"
    if "docx:document:create" in text:
        return "missing_create_scope"
    if "docs:document.media:upload" in text:
        return "missing_media_scope"
    if "permission" in text or "forbidden" in text:
        return "folder_permission_denied"
    return default


def _run_cli(args: list[str], *, cwd: Path | None, timeout: int, command_name: str) -> tuple[dict[str, Any], str, str]:
    env = check_lark_cli_environment()
    command = [env["cli_path"], *args]
    try:
        result = subprocess.run(
            command,
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
            check=False,
            shell=False,
        )
    except subprocess.TimeoutExpired as exc:
        raise FeishuDocumentError("command_timeout", f"{command_name}执行超时", stdout=_redact(exc.stdout or ""), stderr=_redact(exc.stderr or "")) from exc
    except OSError as exc:
        raise FeishuDocumentError("lark_cli_unavailable", f"无法启动lark-cli：{exc}") from exc
    stdout, stderr = _redact(result.stdout), _redact(result.stderr)
    if result.returncode != 0:
        raise FeishuDocumentError(_classify_cli_error(stderr or stdout, f"{command_name}_failed"), f"{command_name}失败（exit={result.returncode}）：{stderr or stdout}", stdout=stdout, stderr=stderr)
    data = extract_last_json_object(result.stdout)
    if data.get("ok") is False or int(data.get("code", 0) or 0) != 0:
        message = str(data.get("message") or data.get("msg") or data.get("error") or data)
        raise FeishuDocumentError(_classify_cli_error(message, f"{command_name}_failed"), f"{command_name}业务失败：{message}", stdout=stdout, stderr=stderr)
    return data, stdout, stderr


def create_feishu_document(title: str, markdown_content: str, parent_token: str | None = None) -> dict[str, Any]:
    token = parent_token or os.getenv("FEISHU_RESULT_FOLDER_TOKEN", "").strip()
    args = [
        "docs", "+create", "--as", "user", "--parent-token", token,
        "--doc-format", "markdown", "--title", title, "--content", markdown_content, "--format", "json",
    ]
    timeout = int(os.getenv("FEISHU_DOC_CREATE_TIMEOUT_SECONDS", "120"))
    try:
        data, stdout, stderr = _run_cli(args, cwd=None, timeout=timeout, command_name="document_create")
    except FeishuDocumentError as exc:
        unsupported_title = "title" in f"{exc.stdout}\n{exc.stderr}".lower() and any(word in f"{exc.stdout}\n{exc.stderr}".lower() for word in ("unknown", "unsupported", "unexpected"))
        if not unsupported_title:
            raise
        compatible_content = f"<title>{html.escape(title)}</title>\n\n{markdown_content}"
        compatible_args = [
            "docs", "+create", "--as", "user", "--parent-token", token,
            "--doc-format", "markdown", "--content", compatible_content, "--format", "json",
        ]
        data, stdout, stderr = _run_cli(compatible_args, cwd=None, timeout=timeout, command_name="document_create")
    document = data.get("data", {}).get("document", {})
    document_id = str(document.get("document_id") or data.get("document_id") or "")
    document_url = str(document.get("url") or data.get("document_url") or data.get("url") or "")
    if not document_id or not document_url:
        raise FeishuDocumentError("cli_output_parse_failed", "创建文档响应缺少document_id或document_url", stdout=stdout, stderr=stderr)
    return {"success": True, "document_id": document_id, "document_url": document_url, "stdout": stdout, "stderr": stderr, "error_type": "", "error_message": ""}


def insert_file_attachment(document_id: str, file_path: str | Path) -> dict[str, Any]:
    path = Path(file_path).resolve()
    if not path.is_file():
        raise FeishuDocumentError("output_file_missing", f"结果文件不存在：{path.name}")
    if path.stat().st_size <= 0:
        raise FeishuDocumentError("invalid_output_file", f"结果文件为空：{path.name}")
    relative_name = f".\\{path.name}" if os.name == "nt" else f"./{path.name}"
    args = ["docs", "+media-insert", "--as", "user", "--doc", document_id, "--file", relative_name, "--type", "file", "--format", "json"]
    data, stdout, stderr = _run_cli(args, cwd=path.parent, timeout=int(os.getenv("FEISHU_DOC_UPLOAD_TIMEOUT_SECONDS", "180")), command_name="attachment_upload")
    payload = data.get("data", {})
    block_id = str(payload.get("block_id") or data.get("block_id") or "")
    file_token = str(payload.get("file_token") or data.get("file_token") or "")
    if not block_id or not file_token:
        raise FeishuDocumentError("cli_output_parse_failed", f"附件响应缺少block_id或file_token：{path.name}", stdout=stdout, stderr=stderr)
    return {"success": True, "document_id": str(payload.get("document_id") or document_id), "block_id": block_id, "file_token": file_token, "file_path": str(path), "stdout": stdout, "stderr": stderr, "error_type": "", "error_message": ""}


def register_final_result_files(task_dir: Path, final_path: Path) -> dict[str, str]:
    tdir = Path(task_dir).resolve()
    output = tdir / "output"
    expected_final = output / FINAL_REVIEW_FILENAME
    if Path(final_path).resolve() != expected_final.resolve():
        raise FeishuDocumentError("wrong_compare_result_file", "第三个附件不是当前任务的最终人工审核结果")
    paths = {
        "full_40": output / OUTPUT_FILENAMES["full_40"],
        "full_51": output / OUTPUT_FILENAMES["full_51"],
        "compare_final": expected_final,
    }
    registered = {key: str(path.relative_to(tdir)) for key, path in paths.items()}
    update_task_meta(tdir, final_result_files=registered)
    return registered


def _registered_files(task_dir: Path, meta: dict[str, Any]) -> dict[str, Path]:
    tdir = Path(task_dir).resolve()
    registered = dict(meta.get("final_result_files") or {})
    expected = {
        "full_40": Path("output") / OUTPUT_FILENAMES["full_40"],
        "full_51": Path("output") / OUTPUT_FILENAMES["full_51"],
        "compare_final": Path("output") / FINAL_REVIEW_FILENAME,
    }
    paths: dict[str, Path] = {}
    for key in ATTACHMENT_ORDER:
        if Path(str(registered.get(key) or "")) != expected[key]:
            raise FeishuDocumentError("wrong_compare_result_file", f"任务未登记正确的{ATTACHMENT_LABELS[key]}路径")
        path = (tdir / expected[key]).resolve()
        if path.parent != (tdir / "output").resolve():
            raise FeishuDocumentError("wrong_compare_result_file", "结果文件路径越界")
        if not path.is_file():
            raise FeishuDocumentError("output_file_missing", f"结果文件不存在：{path.name}")
        if path.stat().st_size <= 0:
            raise FeishuDocumentError("invalid_output_file", f"结果文件为空：{path.name}")
        paths[key] = path
    for key, path in paths.items():
        try:
            workbook = load_workbook(path, read_only=True)
        except Exception as exc:  # noqa: BLE001 - normalize corrupt/non-Excel inputs
            raise FeishuDocumentError("invalid_output_file", f"结果文件不是有效Excel：{path.name}") from exc
        try:
            if key == "compare_final" and any(sheet not in workbook.sheetnames for sheet in SHEET_RULES):
                raise FeishuDocumentError("wrong_compare_result_file", "第三个附件缺少最终人工审核Sheet")
        finally:
            workbook.close()
    return paths


def _beijing_stamp(value: str) -> str:
    try:
        parsed = datetime.fromisoformat(value)
        if not parsed.tzinfo:
            parsed = parsed.replace(tzinfo=timezone.utc)
    except (TypeError, ValueError):
        parsed = datetime.now(timezone.utc)
    return parsed.astimezone(timezone(timedelta(hours=8))).strftime("%Y%m%d_%H%M%S")


def _safe_title(value: str) -> str:
    return re.sub(r"[\x00-\x1f<>]", "_", value).strip()[:200]


def _data_row_count(path: Path) -> int:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return max(0, workbook.active.max_row - 1)
    finally:
        workbook.close()


def _staged_attachment(task_dir: Path, key: str, source: Path) -> Path:
    """Copy one registered result to its stable Feishu display name."""

    if key not in ATTACHMENT_LABELS:
        raise FeishuDocumentError("wrong_compare_result_file", f"未知交付附件：{key}")
    staging_dir = (Path(task_dir).resolve() / "bot" / DELIVERY_STAGING_DIR).resolve()
    expected_parent = (Path(task_dir).resolve() / "bot").resolve()
    if staging_dir.parent != expected_parent:
        raise FeishuDocumentError("wrong_compare_result_file", "飞书交付暂存路径越界")
    staging_dir.mkdir(parents=True, exist_ok=True)
    target = staging_dir / ATTACHMENT_LABELS[key]
    try:
        shutil.copy2(source, target)
    except OSError as exc:
        raise FeishuDocumentError("attachment_stage_failed", f"准备飞书附件失败：{target.name}：{exc}") from exc
    if not target.is_file() or target.stat().st_size != source.stat().st_size:
        raise FeishuDocumentError("attachment_stage_failed", f"准备飞书附件失败：{target.name}")
    return target


def _document_content(task_dir: Path, meta: dict[str, Any], files: dict[str, Path]) -> str:
    review_dir = Path(task_dir) / "review"
    stats = compute_review_stats(load_review_items(review_dir), load_review_state(review_dir))
    fields: list[tuple[str, Any]] = [
        ("任务编号", meta.get("task_id") or Path(task_dir).name),
        ("触发方式", meta.get("trigger_source")),
        ("任务开始时间", meta.get("triggered_at") or meta.get("created_at")),
        ("最终审核确认时间", meta.get("review_completed_at")),
        ("任务完成时间", meta.get("updated_at")),
    ]
    task_lines = "\n".join(f"- {label}：{value}" for label, value in fields if value not in (None, ""))
    input_lines = "\n".join([
        f"- EEA4.0输入文件数量：{int(meta.get('input_40_count') or 0)}",
        f"- EEA5.1输入文件数量：{int(meta.get('input_51_count') or 0)}",
        f"- EEA4.0全量信号数量：{_data_row_count(files['full_40'])}",
        f"- EEA5.1全量信号数量：{_data_row_count(files['full_51'])}",
        f"- 历史版本跳过数量：{int(meta.get('full_compare_skipped_history_count') or 0)}",
    ])
    actual_manual_count = max(0, int(stats.get("manual_confirmed") or 0) - int(stats.get("history_reused") or 0))
    initial_pending_count = int(
        meta.get("initial_pending_manual_count")
        if meta.get("initial_pending_manual_count") is not None
        else int(stats.get("pending_manual") or 0) + actual_manual_count
    )
    result_lines = "\n".join([
        f"- 初始待人工确认数量：{initial_pending_count}",
        f"- 系统判定不同数量：{int(stats.get('system_different') or 0)}",
        f"- 人工确认不同数量：{int(stats.get('manual_different') or 0)}",
        f"- 人工确认相同数量：{int(stats.get('manual_same') or 0)}",
        f"- 历史人工复用数量：{int(stats.get('history_reused') or 0)}",
        f"- 本次人工确认数量：{actual_manual_count}",
        f"- 待人工确认数量：{int(stats.get('pending_manual') or 0)}",
        (
            "- 最终状态：无需新增人工确认，历史人工结论复用完成，最终结果已生成"
            if actual_manual_count == 0
            and int(stats.get("pending_manual") or 0) == 0
            and int(stats.get("history_reused") or 0) > 0
            else (
                "- 最终状态：无需新增人工确认，系统判定完成，最终结果已生成"
                if actual_manual_count == 0 and int(stats.get("pending_manual") or 0) == 0
                else "- 最终状态：人工审核完成，最终结果已生成"
            )
        ),
    ])
    return f"""# EEA4.0与EEA5.1信号差异识别结果

## 一、任务信息
{task_lines}

## 二、输入信息
{input_lines}

## 三、结果统计
{result_lines}

## 四、审核说明
- 信号值描述和单位分别审核；
- 历史人工结果使用严格字段指纹匹配；
- 最终结果以本次任务保存的人工结论为准。

## 五、结果附件
1. {ATTACHMENT_LABELS['full_40']}
2. {ATTACHMENT_LABELS['full_51']}
3. {ATTACHMENT_LABELS['compare_final']}
"""


def _default_delivery(files: dict[str, Path]) -> dict[str, Any]:
    return {
        "status": "not_started", "document_id": "", "document_url": "", "document_title": "",
        "created_at": "", "updated_at": "", "attempt_count": 0, "last_error": "", "error_type": "",
        "attachments": {
            key: {
                "status": "pending",
                "file_path": str(files[key]),
                "display_name": ATTACHMENT_LABELS[key],
                "block_id": "",
                "file_token": "",
                "attempt_count": 0,
                "last_error": "",
                "error_type": "",
            }
            for key in ATTACHMENT_ORDER
        },
        "result_notification": {"status": "pending", "notified_at": "", "attempt_count": 0, "last_error": ""},
    }


def _normalized_delivery(existing: dict[str, Any], files: dict[str, Path]) -> dict[str, Any]:
    delivery = _default_delivery(files)
    delivery.update({key: value for key, value in existing.items() if key not in {"attachments", "result_notification"}})
    for key in ATTACHMENT_ORDER:
        delivery["attachments"][key].update(dict((existing.get("attachments") or {}).get(key) or {}))
    delivery["result_notification"].update(dict(existing.get("result_notification") or {}))
    return delivery


def _save_delivery(task_dir: Path, delivery: dict[str, Any]) -> None:
    delivery["updated_at"] = utc_now_iso()
    status = str(delivery.get("status") or "not_started")
    summary_status = {
        "not_started": "pending",
        "creating": "sending",
        "uploading": "sending",
        "partial_failed": "failed",
        "failed": "failed",
        "ready": "ready",
        "delivered": "delivered",
    }.get(status, status)
    update_task_meta(
        task_dir,
        feishu_delivery=delivery,
        result_delivery_status=summary_status,
        delivery_error=str(delivery.get("last_error") or ""),
    )


def _notify_current_result_status(task_dir: Path) -> None:
    """Best-effort custom-bot notice; delivery errors remain independently retryable."""

    try:
        from .notification_router import notify_result_ready

        notify_result_ready(task_dir)
    except Exception:  # noqa: BLE001 - a notification failure must not replace delivery state
        return


def _acquire_process_claim(task_dir: Path) -> Path:
    path = Path(task_dir) / "bot" / "feishu_doc_delivery.lock"
    path.parent.mkdir(parents=True, exist_ok=True)
    stale_seconds = max(60, int(os.getenv("FEISHU_DOC_LOCK_STALE_SECONDS", "600")))
    if path.exists() and datetime.now().timestamp() - path.stat().st_mtime > stale_seconds:
        path.unlink(missing_ok=True)
    try:
        descriptor = os.open(path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        raise FeishuDocumentError("delivery_already_running", "飞书文档交付正在执行") from exc
    with os.fdopen(descriptor, "w", encoding="utf-8") as handle:
        handle.write(f"pid={os.getpid()} at={utc_now_iso()}")
    return path


def publish_task_result_document(task_dir: str | Path, *, notify: bool = True) -> dict[str, Any]:
    tdir = Path(task_dir).resolve()
    if not _enabled():
        return {"success": False, "error_type": "delivery_disabled", "error_message": "飞书文档交付未启用"}
    try:
        claim = _acquire_process_claim(tdir)
    except FeishuDocumentError as exc:
        return {"success": False, "status": "running", "error_type": exc.error_type, "error_message": str(exc), "last_error": str(exc)}
    try:
        with get_task_lock(tdir):
            meta = load_task_meta(tdir)
            files = _registered_files(tdir, meta)
            delivery = _normalized_delivery(dict(meta.get("feishu_delivery") or {}), files)
            delivery["attempt_count"] = int(delivery.get("attempt_count") or 0) + 1
            delivery["status"] = "creating" if not delivery.get("document_id") else "uploading"
            delivery["last_error"] = ""
            _save_delivery(tdir, delivery)
        if not delivery.get("document_id"):
            title = delivery.get("document_title") or _safe_title(
                f"{_beijing_stamp(str(meta.get('review_completed_at') or ''))}_信号矩阵全量对比最终结果"
            )
            created = create_feishu_document(title, _document_content(tdir, meta, files))
            delivery.update({
                "document_id": created["document_id"], "document_url": created["document_url"],
                "document_title": title, "created_at": utc_now_iso(), "status": "uploading",
            })
            _save_delivery(tdir, delivery)
        for key in ATTACHMENT_ORDER:
            attachment = delivery["attachments"][key]
            if attachment.get("status") == "success":
                continue
            attachment["status"] = "uploading"
            attachment["attempt_count"] = int(attachment.get("attempt_count") or 0) + 1
            _save_delivery(tdir, delivery)
            try:
                upload_path = _staged_attachment(tdir, key, files[key])
                uploaded = insert_file_attachment(str(delivery["document_id"]), upload_path)
            except FeishuDocumentError as exc:
                attachment.update({"status": "failed", "last_error": str(exc), "error_type": exc.error_type, "last_stdout": exc.stdout, "last_stderr": exc.stderr})
                delivery.update({"status": "partial_failed", "last_error": str(exc), "error_type": exc.error_type, "last_stdout": exc.stdout, "last_stderr": exc.stderr})
                _save_delivery(tdir, delivery)
                if notify:
                    _notify_current_result_status(tdir)
                return {"success": False, **delivery}
            attachment.update({
                "status": "success",
                "display_name": ATTACHMENT_LABELS[key],
                "block_id": uploaded["block_id"],
                "file_token": uploaded["file_token"],
                "last_error": "",
                "error_type": "",
            })
            _save_delivery(tdir, delivery)
        delivery.update({"status": "ready", "last_error": "", "error_type": ""})
        _save_delivery(tdir, delivery)
        if notify:
            retry_result_notification(tdir)
            delivery = dict(load_task_meta(tdir).get("feishu_delivery") or delivery)
        return {"success": delivery.get("status") in {"ready", "delivered"}, **delivery}
    except FeishuDocumentError as exc:
        meta = load_task_meta(tdir)
        delivery = dict(meta.get("feishu_delivery") or {})
        delivery.update({"status": "failed", "last_error": str(exc), "error_type": exc.error_type, "last_stdout": exc.stdout, "last_stderr": exc.stderr})
        _save_delivery(tdir, delivery)
        if notify:
            _notify_current_result_status(tdir)
        return {"success": False, **delivery}
    finally:
        claim.unlink(missing_ok=True)


def retry_failed_attachments(task_id_or_dir: str | Path) -> dict[str, Any]:
    tdir = Path(task_id_or_dir)
    if not tdir.exists():
        from .bot_task_store import task_dir
        tdir = task_dir(str(task_id_or_dir))
    meta = load_task_meta(tdir)
    if not (meta.get("feishu_delivery") or {}).get("document_id"):
        raise FeishuDocumentError("document_create_failed", "尚未创建飞书文档，不能只重试附件")
    return publish_task_result_document(tdir, notify=True)


def retry_result_notification(task_id_or_dir: str | Path, *, custom_client: Any | None = None, force: bool = False) -> bool:
    tdir = Path(task_id_or_dir)
    if not tdir.exists():
        from .bot_task_store import task_dir
        tdir = task_dir(str(task_id_or_dir))
    meta = load_task_meta(tdir)
    delivery = dict(meta.get("feishu_delivery") or {})
    if delivery.get("status") not in {"ready", "delivered"} or not delivery.get("document_url"):
        return False
    notice = dict(delivery.get("result_notification") or {})
    if notice.get("status") == "sent" and not force:
        return True
    notice.update({"status": "sending", "attempt_count": int(notice.get("attempt_count") or 0) + 1, "last_error": ""})
    delivery["result_notification"] = notice
    _save_delivery(tdir, delivery)
    from .notification_router import notify_result_ready
    ok = notify_result_ready(tdir, custom_client=custom_client, force=force)
    if ok:
        notice.update({"status": "sent", "notified_at": utc_now_iso(), "last_error": ""})
        delivery["status"] = "delivered"
        update_task_meta(tdir, feishu_delivery=delivery, status="delivered", result_delivery_status="delivered", result_delivered_at=utc_now_iso())
    else:
        notice.update({"status": "failed", "last_error": "群机器人最终通知发送失败"})
        _save_delivery(tdir, delivery)
    return ok
