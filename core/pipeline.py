"""Subprocess wrapper around the legacy EEA 4.0/5.1 matrix scripts.

The legacy scripts intentionally remain the source of business rules.  This
module prepares an isolated task directory, copies the scripts there, executes
01 -> 02 -> 03 with cwd set to that task directory, then summarizes outputs for
Streamlit.
"""

from __future__ import annotations

import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any


REPO_ROOT = Path(__file__).resolve().parents[1]
LEGACY_SCRIPTS = [
    "common_matrix_utils_local_v13.py",
    "01_extract_full_matrix_local_v13.py",
    "02_generate_dedup_signals_local_v13.py",
    "03_generate_compare_file5_local_v13.py",
]
EXTRACT_SCRIPT = "01_extract_full_matrix_local_v13.py"
DEDUP_SCRIPT = "02_generate_dedup_signals_local_v13.py"
COMPARE_SCRIPT = "03_generate_compare_file5_local_v13.py"

OUTPUT_FILENAMES = {
    "full_40": "26R1 4.0全量信号矩阵清单.xlsx",
    "full_51": "26R2 5.1全量信号矩阵清单.xlsx",
    "dedup_40": "26R1 4.0全量信号-同名去重后.xlsx",
    "dedup_51": "26R2 5.1全量信号-同名去重后.xlsx",
    "compare": "4.0和5.1同一信号差异点识别.xlsx",
}


@dataclass
class PipelineStepResult:
    script: str
    returncode: int
    stdout: str
    stderr: str


class PipelineStepError(RuntimeError):
    """Raised when a legacy script exits with a non-zero return code."""

    def __init__(self, result: PipelineStepResult) -> None:
        self.result = result
        super().__init__(
            f"Legacy script failed: {result.script}\n"
            f"returncode={result.returncode}\n"
            f"stdout:\n{result.stdout}\n"
            f"stderr:\n{result.stderr}"
        )


def _as_path(path: str | Path) -> Path:
    return Path(path).expanduser().resolve()


def _task_dir_from_output(output_dir: str | Path) -> Path:
    return _as_path(output_dir).parent


def _copy_dir_contents(src: Path, dst: Path) -> None:
    dst.mkdir(parents=True, exist_ok=True)
    if src == dst:
        return
    for item in src.iterdir():
        target = dst / item.name
        if item.is_dir():
            if target.exists():
                shutil.rmtree(target)
            shutil.copytree(item, target)
        else:
            shutil.copy2(item, target)


def ensure_legacy_scripts(task_dir: str | Path) -> list[Path]:
    """Copy required legacy scripts into *task_dir* and return copied paths."""

    task_path = _as_path(task_dir)
    task_path.mkdir(parents=True, exist_ok=True)
    missing = [name for name in LEGACY_SCRIPTS if not (REPO_ROOT / name).is_file()]
    if missing:
        raise FileNotFoundError("缺少 legacy 脚本：" + ", ".join(missing))

    copied: list[Path] = []
    for name in LEGACY_SCRIPTS:
        src = REPO_ROOT / name
        dst = task_path / name
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        copied.append(dst)
    return copied


def _prepare_task_dirs(input_40_dir: str | Path, input_51_dir: str | Path, output_dir: str | Path) -> Path:
    output_path = _as_path(output_dir)
    task_dir = output_path.parent
    input_40_path = _as_path(input_40_dir)
    input_51_path = _as_path(input_51_dir)

    (task_dir / "input" / "4.0").mkdir(parents=True, exist_ok=True)
    (task_dir / "input" / "5.1").mkdir(parents=True, exist_ok=True)
    output_path.mkdir(parents=True, exist_ok=True)

    _copy_dir_contents(input_40_path, task_dir / "input" / "4.0")
    _copy_dir_contents(input_51_path, task_dir / "input" / "5.1")
    ensure_legacy_scripts(task_dir)
    return task_dir


def _run_script(task_dir: Path, script_name: str) -> PipelineStepResult:
    proc = subprocess.run(
        [sys.executable, script_name],
        cwd=task_dir,
        text=True,
        capture_output=True,
        check=False,
    )
    result = PipelineStepResult(script_name, proc.returncode, proc.stdout, proc.stderr)
    if proc.returncode != 0:
        raise PipelineStepError(result)
    return result


def _count_rows(path: Path, sheet_name: str | None = None) -> int:
    if not path.exists():
        return 0
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    count = max((ws.max_row or 1) - 1, 0)
    wb.close()
    return count


def _parse_log_int(log_path: Path, label: str) -> int:
    if not log_path.exists():
        return 0
    text = log_path.read_text(encoding="utf-8", errors="ignore")
    match = re.search(re.escape(label) + r"[:：]\s*(\d+)", text)
    return int(match.group(1)) if match else 0


def collect_statistics(output_dir: str | Path) -> dict[str, int]:
    output_path = _as_path(output_dir)
    compare_log = output_path / "03_generate_compare_file5_local_v13_log.txt"
    compare_file = output_path / OUTPUT_FILENAMES["compare"]
    return {
        "4.0 全量信号数量": _count_rows(output_path / OUTPUT_FILENAMES["full_40"]),
        "5.1 全量信号数量": _count_rows(output_path / OUTPUT_FILENAMES["full_51"]),
        "4.0 去重后信号数量": _count_rows(output_path / OUTPUT_FILENAMES["dedup_40"]),
        "5.1 去重后信号数量": _count_rows(output_path / OUTPUT_FILENAMES["dedup_51"]),
        "完全同名匹配信号数": _parse_log_int(compare_log, "完全同名匹配信号数"),
        "sheet1 差异行数": _count_rows(compare_file, "完全同名匹配对比结果"),
        "sheet2 vcu-hcu 差异行数": _count_rows(compare_file, "vcu-hcu 同名匹配"),
    }


def _output_files(output_dir: Path) -> dict[str, Path]:
    return {key: output_dir / filename for key, filename in OUTPUT_FILENAMES.items()}


def run_extract(input_40_dir: str | Path, input_51_dir: str | Path, output_dir: str | Path) -> PipelineStepResult:
    task_dir = _prepare_task_dirs(input_40_dir, input_51_dir, output_dir)
    return _run_script(task_dir, EXTRACT_SCRIPT)


def run_dedup(output_dir: str | Path) -> PipelineStepResult:
    output_path = _as_path(output_dir)
    task_dir = _task_dir_from_output(output_path)
    ensure_legacy_scripts(task_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    return _run_script(task_dir, DEDUP_SCRIPT)


def run_compare(output_dir: str | Path) -> PipelineStepResult:
    output_path = _as_path(output_dir)
    task_dir = _task_dir_from_output(output_path)
    ensure_legacy_scripts(task_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    return _run_script(task_dir, COMPARE_SCRIPT)


def run_all(input_40_dir: str | Path, input_51_dir: str | Path, output_dir: str | Path) -> dict[str, Any]:
    output_path = _as_path(output_dir)
    logs: list[PipelineStepResult] = []
    logs.append(run_extract(input_40_dir, input_51_dir, output_path))
    logs.append(run_dedup(output_path))
    logs.append(run_compare(output_path))

    files = _output_files(output_path)
    missing = [str(path) for path in files.values() if not path.exists()]
    if missing:
        raise FileNotFoundError("流程完成但缺少输出文件：" + ", ".join(missing))

    return {
        "output_dir": output_path,
        "files": files,
        "statistics": collect_statistics(output_path),
        "logs": [result.__dict__ for result in logs],
    }
