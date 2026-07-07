# -*- coding: utf-8 -*-
"""
03_generate_compare_file5_local_v13.py

基于文件三、文件四生成文件五：
《4.0和5.1同一信号差异点识别.xlsx》

说明：
ECU 收发状态只作为追溯字段带出，不参与差异筛选。
标准化状态已改为小写 r/s。
"""

from __future__ import annotations

import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common_matrix_utils_local_v13 import (
    cell_text,
    norm_header,
    values_equal,
    strip_vcu_hcu_prefix,
)


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"

DEDUP_40_FILE = OUTPUT_DIR / "26R1 4.0全量信号-同名去重后.xlsx"
DEDUP_51_FILE = OUTPUT_DIR / "26R2 5.1全量信号-同名去重后.xlsx"

OUTPUT_FILE = OUTPUT_DIR / "4.0和5.1同一信号差异点识别.xlsx"
LOG_FILE = OUTPUT_DIR / "03_generate_compare_file5_local_v13_log.txt"


COMPARE_FIELDS = [
    ("bit_length", "信号长度"),
    ("resolution", "精度"),
    ("offset", "偏移量"),
    ("signal_min", "物理最小值"),
    ("signal_max", "物理最大值"),
    ("unit", "单位"),
    ("signal_value_description", "信号值描述"),
]

EXTRA_FIELDS = [
    ("source_files", "信号来源文件"),
    ("ecu_status_raw", "ECU收发状态_原始"),
    ("ecu_status_std", "ECU收发状态_标准化"),
    ("send_ecu_summary", "发送ECU汇总"),
    ("receive_ecu_summary", "接收ECU汇总"),
]

COLUMN_CANDIDATES = {
    "signal_name": ["信号名称", "Signal Name", "信号名"],
    "bit_length": ["信号长度", "Bit Length", "Signal Size"],
    "resolution": ["精度", "Resolution", "Factor"],
    "offset": ["偏移量", "Offset"],
    "signal_min": ["物理最小值", "Signal Min", "Minimum", "Min"],
    "signal_max": ["物理最大值", "Signal Max", "Maximum", "Max"],
    "unit": ["单位", "Unit"],
    "signal_value_description": ["信号值描述", "Signal Value Description", "Value Description"],
    "source_files": ["信号来源文件"],
    "ecu_status_raw": ["ECU收发状态_原始", "ECU接收和发送状态"],
    "ecu_status_std": ["ECU收发状态_标准化"],
    "send_ecu_summary": ["发送ECU汇总"],
    "receive_ecu_summary": ["接收ECU汇总"],
}


def match_col(value: Any, names: List[str]) -> bool:
    raw = cell_text(value)
    compact = norm_header(raw)
    if not raw:
        return False
    for name in names:
        n = norm_header(name)
        if compact == n or n in compact:
            return True
    return False


def find_columns(ws) -> Tuple[int, Dict[str, int]]:
    best_row = 1
    best_map: Dict[str, int] = {}
    best_score = -1

    for row_idx in range(1, min(ws.max_row or 0, 10) + 1):
        cmap: Dict[str, int] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            for code, names in COLUMN_CANDIDATES.items():
                if code not in cmap and match_col(v, names):
                    cmap[code] = col_idx
                    break
        if "signal_name" in cmap and len(cmap) > best_score:
            best_row = row_idx
            best_map = cmap
            best_score = len(cmap)

    if "signal_name" not in best_map:
        raise ValueError(f"未找到信号名称列：{ws.title}")

    return best_row, best_map


def read_dedup(input_file: Path) -> Dict[str, Dict[str, str]]:
    wb = load_workbook(input_file, data_only=True, read_only=False)
    ws = wb.active
    header_row, cmap = find_columns(ws)

    records: Dict[str, Dict[str, str]] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = cell_text(ws.cell(row=row_idx, column=cmap["signal_name"]).value)
        if not name:
            continue

        rec: Dict[str, str] = {"signal_name": name}
        for code, _title in COMPARE_FIELDS + EXTRA_FIELDS:
            col = cmap.get(code)
            rec[code] = cell_text(ws.cell(row=row_idx, column=col).value) if col else ""

        if name not in records:
            records[name] = rec

    wb.close()
    return records


def build_diff_list(r40: Dict[str, str], r51: Dict[str, str]) -> str:
    lines: List[str] = []
    for code, title in COMPARE_FIELDS:
        v40 = r40.get(code, "")
        v51 = r51.get(code, "")
        if not values_equal(code, v40, v51):
            lines.append(f"{title}：4.0={v40}；5.1={v51}")
    return "\n".join(lines)


def build_headers(include_match_key: bool = False) -> List[str]:
    headers = ["4.0信号名", "5.1信号名"]
    if include_match_key:
        headers.append("去前缀后匹配名")
    headers.append("差异点list")

    for _code, title in COMPARE_FIELDS:
        headers.append(f"4.0_{title}")
        headers.append(f"5.1_{title}")

    for _code, title in EXTRA_FIELDS:
        headers.append(f"4.0_{title}")
        headers.append(f"5.1_{title}")

    return headers


def build_row(r40: Dict[str, str], r51: Dict[str, str], diff_list: str, match_key: str = "") -> List[str]:
    row = [r40["signal_name"], r51["signal_name"]]
    if match_key:
        row.append(match_key)
    row.append(diff_list)

    for code, _title in COMPARE_FIELDS:
        row.append(r40.get(code, ""))
        row.append(r51.get(code, ""))

    for code, _title in EXTRA_FIELDS:
        row.append(r40.get(code, ""))
        row.append(r51.get(code, ""))

    return row


def build_exact_match_rows(rec40: Dict[str, Dict[str, str]], rec51: Dict[str, Dict[str, str]]) -> Tuple[List[List[str]], set, set]:
    rows: List[List[str]] = []
    matched40 = set()
    matched51 = set()

    for name in sorted(set(rec40.keys()) & set(rec51.keys())):
        r40 = rec40[name]
        r51 = rec51[name]
        matched40.add(name)
        matched51.add(name)

        diff = build_diff_list(r40, r51)
        if diff:
            rows.append(build_row(r40, r51, diff))

    return rows, matched40, matched51


def build_vcu_hcu_rows(
    rec40: Dict[str, Dict[str, str]],
    rec51: Dict[str, Dict[str, str]],
    matched40: set,
    matched51: set,
) -> List[List[str]]:
    remain40 = {
        name: rec
        for name, rec in rec40.items()
        if name not in matched40 and name.upper().startswith(("VCU", "HCU"))
    }
    remain51 = {
        name: rec
        for name, rec in rec51.items()
        if name not in matched51 and name.upper().startswith(("VCU", "HCU"))
    }

    idx51: Dict[str, List[str]] = {}
    for name in remain51:
        key = strip_vcu_hcu_prefix(name)
        idx51.setdefault(key, []).append(name)

    rows: List[List[str]] = []
    used51 = set()

    for name40 in sorted(remain40.keys()):
        key = strip_vcu_hcu_prefix(name40)
        candidates = [x for x in idx51.get(key, []) if x not in used51]
        if not candidates:
            continue

        name51 = candidates[0]
        used51.add(name51)

        r40 = remain40[name40]
        r51 = remain51[name51]
        diff = build_diff_list(r40, r51)

        if diff:
            rows.append(build_row(r40, r51, diff, match_key=key))

    return rows


def write_sheet(ws, headers: List[str], rows: List[List[str]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in range(1, ws.max_column + 1):
        header = cell_text(ws.cell(row=1, column=col).value)
        letter = get_column_letter(col)
        if "描述" in header or "差异点" in header or "ECU" in header or "来源" in header:
            ws.column_dimensions[letter].width = 95
        elif "信号名" in header:
            ws.column_dimensions[letter].width = 36
        else:
            ws.column_dimensions[letter].width = 16


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    rec40 = read_dedup(DEDUP_40_FILE)
    rec51 = read_dedup(DEDUP_51_FILE)

    sheet1_rows, matched40, matched51 = build_exact_match_rows(rec40, rec51)
    sheet2_rows = build_vcu_hcu_rows(rec40, rec51, matched40, matched51)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "完全同名匹配对比结果"
    write_sheet(ws1, build_headers(include_match_key=False), sheet1_rows)

    ws2 = wb.create_sheet("vcu-hcu 同名匹配")
    write_sheet(ws2, build_headers(include_match_key=True), sheet2_rows)

    wb.save(OUTPUT_FILE)
    wb.close()

    logs = [
        f"4.0去重信号数：{len(rec40)}",
        f"5.1去重信号数：{len(rec51)}",
        f"完全同名匹配信号数：{len(matched40)}",
        f"sheet1存在差异行数：{len(sheet1_rows)}",
        f"sheet2存在差异行数：{len(sheet2_rows)}",
        f"输出文件：{OUTPUT_FILE}",
    ]

    LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
    print("\n".join(logs), flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("脚本运行异常：")
        print(traceback.format_exc())
        sys.exit(1)
