from __future__ import annotations

from pathlib import Path
from uuid import uuid4
import io
import zipfile

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from core import run_all
from core.pipeline import (
    COMPARE_FILE,
    DEDUP_40_FILE,
    DEDUP_51_FILE,
    FULL_40_FILE,
    FULL_51_FILE,
)

OUTPUT_FILES = [FULL_40_FILE, FULL_51_FILE, DEDUP_40_FILE, DEDUP_51_FILE, COMPARE_FILE]


def _save_uploads(uploaded_files, target_dir: Path) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    for uploaded in uploaded_files:
        safe_name = Path(uploaded.name).name
        (target_dir / safe_name).write_bytes(uploaded.getbuffer())


def _excel_data_rows(path: Path, sheet_name: str | int = 0) -> int:
    if not path.exists():
        return 0
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if isinstance(sheet_name, str) else workbook.worksheets[sheet_name]
        return max(worksheet.max_row - 1, 0)
    finally:
        workbook.close()


def _build_statistics(output_dir: Path) -> dict[str, int]:
    compare_path = output_dir / COMPARE_FILE
    stats = {
        "4.0 全量信号数量": _excel_data_rows(output_dir / FULL_40_FILE),
        "5.1 全量信号数量": _excel_data_rows(output_dir / FULL_51_FILE),
        "4.0 去重后信号数量": _excel_data_rows(output_dir / DEDUP_40_FILE),
        "5.1 去重后信号数量": _excel_data_rows(output_dir / DEDUP_51_FILE),
        "完全同名匹配信号数": _excel_data_rows(compare_path, 0),
        "sheet1 差异行数": _excel_data_rows(compare_path, 0),
        "sheet2 vcu-hcu 差异行数": _excel_data_rows(compare_path, 1) if compare_path.exists() else 0,
    }
    return stats


def _zip_outputs(output_dir: Path) -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for file_name in OUTPUT_FILES:
            file_path = output_dir / file_name
            if file_path.exists():
                archive.write(file_path, arcname=file_name)
    buffer.seek(0)
    return buffer.getvalue()


def _download_results(output_dir: Path) -> None:
    st.subheader("结果下载")
    for file_name in OUTPUT_FILES:
        file_path = output_dir / file_name
        if file_path.exists():
            st.download_button(
                label=f"下载 {file_name}",
                data=file_path.read_bytes(),
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.warning(f"未找到结果文件：{file_name}")

    st.download_button(
        label="下载全部结果 zip",
        data=_zip_outputs(output_dir),
        file_name="matrix-signal-diff-results.zip",
        mime="application/zip",
    )


def main() -> None:
    st.set_page_config(page_title="EEA 4.0/5.1 矩阵同一信号差异识别工具", layout="wide")
    st.title("EEA 4.0/5.1 矩阵同一信号差异识别工具")
    st.caption("本地 Demo：上传 4.0 与 5.1 矩阵 Excel 文件后，调用 core.run_all() 完成抽取、去重和差异识别。")

    uploaded_40 = st.file_uploader("上传 EEA 4.0 矩阵文件（xlsx/xlsm，可多选）", type=["xlsx", "xlsm"], accept_multiple_files=True)
    uploaded_51 = st.file_uploader("上传 EEA 5.1 矩阵文件（xlsx/xlsm，可多选）", type=["xlsx", "xlsm"], accept_multiple_files=True)

    if "last_output_dir" in st.session_state:
        _download_results(Path(st.session_state["last_output_dir"]))

    if st.button("开始识别", type="primary"):
        if not uploaded_40 or not uploaded_51:
            st.error("请同时上传 EEA 4.0 和 EEA 5.1 矩阵文件。")
            return

        task_id = uuid4().hex
        task_root = Path("temp") / task_id
        input_40_dir = task_root / "input" / "4.0"
        input_51_dir = task_root / "input" / "5.1"
        output_dir = task_root / "output"

        progress = st.progress(0, text="创建临时任务目录...")
        try:
            input_40_dir.mkdir(parents=True, exist_ok=True)
            input_51_dir.mkdir(parents=True, exist_ok=True)
            output_dir.mkdir(parents=True, exist_ok=True)

            progress.progress(20, text="保存上传文件...")
            _save_uploads(uploaded_40, input_40_dir)
            _save_uploads(uploaded_51, input_51_dir)

            progress.progress(40, text="执行矩阵抽取、去重和差异识别...")
            run_all(input_40_dir, input_51_dir, output_dir)

            progress.progress(85, text="读取统计信息...")
            stats = _build_statistics(output_dir)

            progress.progress(100, text="处理完成。")
            st.success(f"识别完成，任务 ID：{task_id}")
            st.session_state["last_output_dir"] = str(output_dir)

            st.subheader("基础统计")
            stats_df = pd.DataFrame([{"指标": key, "数量": value} for key, value in stats.items()])
            st.dataframe(stats_df, hide_index=True, use_container_width=True)

            _download_results(output_dir)
        except Exception as exc:  # Streamlit should show clear processing failures.
            progress.empty()
            st.error("处理失败，请检查上传文件格式、矩阵 sheet、依赖安装和原始业务脚本是否存在。")
            st.exception(exc)


if __name__ == "__main__":
    main()
