# -*- coding: utf-8 -*-
"""
01_extract_full_matrix_local_v13.py

功能：
从本地 input\\4.0 和 input\\5.1 读取所有 xlsx/xlsm 矩阵文件，
只提取真正的信号矩阵 sheet，生成文件一、文件二。

本版口径：
1. 只收集通过核心表头校验的矩阵 sheet，避免收集 History 等非矩阵 sheet。
2. ECU 收发状态列只从信号定义字段右侧识别。
3. 标准化状态改为小写：
   R/r -> r，接收
   T/t/S/s -> s，发送
4. 4.0 的 ECU 状态输出采用“第一个出现的表头为主，后续放括号”的格式：
   示例：
   TCU_PHEV:R(HighRegulationArea:R,LowRegulationArea:R)
   标准化列：
   TCU_PHEV:r(HighRegulationArea:r,LowRegulationArea:r)
5. 5.1 不做括号合并，保持每个 ECU 一行：
   VCU_5IC_DK:s
   VCU_5IH_DK:s

输出：
output\\26R1 4.0全量信号矩阵清单.xlsx
output\\26R2 5.1全量信号矩阵清单.xlsx

运行：
cd /d D:\\signal_compare
python 01_extract_full_matrix_local_v13.py
"""

from __future__ import annotations

import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common_matrix_utils_local_v13 import (
    cell_text,
    norm_header,
    is_excel_file,
    normalize_ecu_state,
    is_status_value,
    join_lines,
)


BASE_DIR = Path(__file__).resolve().parent
INPUT_40_DIR = BASE_DIR / "input" / "4.0"
INPUT_51_DIR = BASE_DIR / "input" / "5.1"
OUTPUT_DIR = BASE_DIR / "output"

OUTPUT_40_FILE = OUTPUT_DIR / "26R1 4.0全量信号矩阵清单.xlsx"
OUTPUT_51_FILE = OUTPUT_DIR / "26R2 5.1全量信号矩阵清单.xlsx"
LOG_FILE = OUTPUT_DIR / "01_extract_full_matrix_local_v13_log.txt"


HEADER_ALIASES = {
    "msg_id": ["Msg ID", "Message ID", "ID", "报文ID", "报文标识符"],
    "frame_format": ["Frame Format", "CAN Type", "帧格式"],
    "msg_name": ["Msg Name", "Message Name", "报文名称"],
    "msg_length": ["Msg Length", "Msg Length(Byte)", "Msg Size", "DLC", "报文长度"],
    "msg_send_type": ["Msg Send Type", "报文发送类型", "Send Type"],
    "transmitter": ["Transmitter", "Sender", "发送节点", "发送方"],
    "signal_name": ["Signal Name", "Signal Name 信号名称", "Signal", "信号名称", "信号名"],
    "start_bit": ["Start Bit", "StartBit", "Start Bit(LSB)", "起始位"],
    "bit_length": ["Bit Length", "Signal Size", "Signal Length", "信号长度", "长度"],
    "data_type": ["Data Type", "Value Type", "数据类型"],
    "resolution": ["Resolution", "Factor", "Scale", "精度", "分辨率"],
    "offset": ["Offset", "偏移量", "偏移"],
    "signal_min": ["Signal Min. Value(phys)", "Signal Min. Value", "Signal Min", "Minimum", "Min", "物理最小值"],
    "signal_max": ["Signal Max. Value(phys)", "Signal Max. Value", "Signal Max", "Maximum", "Max", "物理最大值"],
    "unit": ["Unit", "单位"],
    "signal_value_description": ["Signal Value Description", "Value Description", "Value Table", "信号值描述", "信号值"],
    "comment": ["Comment", "Signal Description", "Description", "备注", "描述"],
}

STRICT_MATRIX_REQUIRED_FIELDS = [
    "msg_id",
    "frame_format",
    "msg_name",
    "msg_length",
    "msg_send_type",
    "transmitter",
    "signal_name",
    "start_bit",
    "bit_length",
    "data_type",
    "resolution",
    "offset",
    "signal_min",
    "signal_max",
]

MAX_MISSING_REQUIRED_FIELDS = 1

OUTPUT_COLUMNS = [
    ("arch_version", "架构版本"),
    ("source_file", "来源文件"),
    ("source_sheet", "来源Sheet"),
    ("source_row", "来源行号"),
    ("msg_id", "Msg ID"),
    ("frame_format", "Frame Format"),
    ("msg_name", "Msg Name"),
    ("msg_length", "Msg Length(Byte)"),
    ("msg_send_type", "Msg Send Type"),
    ("transmitter", "Transmitter"),
    ("signal_name", "Signal Name 信号名称"),
    ("start_bit", "Start Bit"),
    ("bit_length", "信号长度"),
    ("data_type", "Data Type"),
    ("resolution", "精度"),
    ("offset", "偏移量"),
    ("signal_min", "物理最小值"),
    ("signal_max", "物理最大值"),
    ("unit", "单位"),
    ("signal_value_description", "信号值描述"),
    ("comment", "备注"),
    ("ecu_status_raw", "ECU收发状态_原始"),
    ("ecu_status_std", "ECU收发状态_标准化"),
    ("send_ecu_summary", "发送ECU汇总"),
    ("receive_ecu_summary", "接收ECU汇总"),
]


def log(msg: str, logs: List[str]) -> None:
    print(msg, flush=True)
    logs.append(msg)


def header_match(value: Any, aliases: List[str]) -> bool:
    raw = cell_text(value)
    compact = norm_header(raw)
    if not raw:
        return False
    for alias in aliases:
        a = norm_header(alias)
        if not a:
            continue
        if compact == a or a in compact:
            return True
    return False


def scan_header_row(ws, row_idx: int) -> Dict[str, int]:
    field_map: Dict[str, int] = {}
    max_col = ws.max_column or 0

    for col_idx in range(1, max_col + 1):
        v = ws.cell(row=row_idx, column=col_idx).value
        for code, aliases in HEADER_ALIASES.items():
            if code not in field_map and header_match(v, aliases):
                field_map[code] = col_idx
                break

    return field_map


def is_valid_matrix_header(field_map: Dict[str, int]) -> Tuple[bool, List[str]]:
    missing = [code for code in STRICT_MATRIX_REQUIRED_FIELDS if code not in field_map]

    must_have = ["signal_name", "bit_length", "resolution", "offset", "signal_min", "signal_max"]
    if any(code not in field_map for code in must_have):
        return False, missing

    ok = len(missing) <= MAX_MISSING_REQUIRED_FIELDS
    return ok, missing


def find_header_row_and_map(ws, max_scan_rows: int = 40) -> Tuple[int, Dict[str, int], List[str]]:
    best_row = 0
    best_map: Dict[str, int] = {}
    best_missing: List[str] = []
    best_score = -1

    max_row = min(ws.max_row or 0, max_scan_rows)

    for row_idx in range(1, max_row + 1):
        field_map = scan_header_row(ws, row_idx)
        ok, missing = is_valid_matrix_header(field_map)
        score = len([c for c in STRICT_MATRIX_REQUIRED_FIELDS if c in field_map])

        if ok and score > best_score:
            best_row = row_idx
            best_map = field_map
            best_missing = missing
            best_score = score

    return best_row, best_map, best_missing


def is_signal_row_name(name: str) -> bool:
    name = cell_text(name)
    if not name:
        return False
    n = norm_header(name)
    invalid = {
        "signalname",
        "signal",
        "signalname信号名称",
        "信号名称",
        "信号名",
    }
    if n in invalid:
        return False
    return True


def detect_ecu_status_columns(ws, header_row: int, field_map: Dict[str, int]) -> List[Tuple[int, str]]:
    signal_col = field_map["signal_name"]
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    # 从信号定义字段右侧开始扫
    right_start_col = max(field_map.values()) + 1

    signal_rows: List[int] = []
    for row_idx in range(header_row + 1, max_row + 1):
        sig = cell_text(ws.cell(row=row_idx, column=signal_col).value)
        if is_signal_row_name(sig):
            signal_rows.append(row_idx)
        if len(signal_rows) >= 1000:
            break

    if not signal_rows:
        return []

    candidates: List[Tuple[int, str]] = []

    for col_idx in range(right_start_col, max_col + 1):
        header = cell_text(ws.cell(row=header_row, column=col_idx).value)
        if not header:
            continue

        header_norm = norm_header(header)
        exclude_keywords = [
            "asil", "safety", "安全", "comment", "remark", "备注", "description", "描述",
            "history", "version", "版本", "date", "time", "修改", "issue",
        ]
        if any(k in header_norm for k in exclude_keywords):
            continue

        status_count = 0
        nonempty_count = 0
        invalid_nonempty_count = 0

        for row_idx in signal_rows:
            v = cell_text(ws.cell(row=row_idx, column=col_idx).value)
            if not v:
                continue
            nonempty_count += 1
            if is_status_value(v):
                status_count += 1
            else:
                invalid_nonempty_count += 1

        if status_count == 0:
            continue

        if nonempty_count > 0 and status_count / nonempty_count < 0.6:
            continue

        candidates.append((col_idx, header))

    return candidates


def format_grouped_status(pairs: List[Tuple[str, str]]) -> str:
    """
    4.0 专用：
    第一个 ECU 为主，后续 ECU 放括号。
    输入示例：
    [("TCU_PHEV", "R"), ("HighRegulationArea", "R"), ("LowRegulationArea", "R")]
    输出：
    TCU_PHEV:R(HighRegulationArea:R,LowRegulationArea:R)
    """
    if not pairs:
        return ""

    first_ecu, first_state = pairs[0]
    rest = pairs[1:]

    first_part = f"{first_ecu}:{first_state}"
    if not rest:
        return first_part

    rest_part = ",".join(f"{ecu}:{state}" for ecu, state in rest)
    return f"{first_part}({rest_part})"


def format_grouped_names(names: List[str]) -> str:
    """
    4.0 发送/接收 ECU 汇总专用：
    第一个 ECU 为主，后续 ECU 放括号，不带 r/s。
    例如：
    ["EMS_PHEV", "HighRegulationArea", "LowRegulationArea"]
    -> EMS_PHEV(HighRegulationArea,LowRegulationArea)
    """
    names = [cell_text(x) for x in names if cell_text(x)]
    if not names:
        return ""
    first = names[0]
    rest = names[1:]
    if not rest:
        return first
    return f"{first}({','.join(rest)})"


def format_plain_status_lines(pairs: List[Tuple[str, str]]) -> str:
    """
    5.1 保持原样：每个 ECU 一行。
    """
    return "\n".join(f"{ecu}:{state}" for ecu, state in pairs)


def build_ecu_status_for_row(ws, row_idx: int, ecu_cols: List[Tuple[int, str]], arch_version: str) -> Dict[str, str]:
    raw_pairs: List[Tuple[str, str]] = []
    std_pairs: List[Tuple[str, str]] = []
    send: List[str] = []
    recv: List[str] = []

    for col_idx, ecu_name in ecu_cols:
        raw = cell_text(ws.cell(row=row_idx, column=col_idx).value)
        std = normalize_ecu_state(raw)

        if not std or std == "x":
            continue

        raw_pairs.append((ecu_name, raw))
        std_pairs.append((ecu_name, std))

        if std == "s":
            send.append(ecu_name)
        elif std == "r":
            recv.append(ecu_name)

    if arch_version == "4.0":
        raw_text = format_grouped_status(raw_pairs)
        std_text = format_grouped_status(std_pairs)
        send_summary = format_grouped_names(send)
        receive_summary = format_grouped_names(recv)
    else:
        raw_text = format_plain_status_lines(raw_pairs)
        std_text = format_plain_status_lines(std_pairs)
        send_summary = join_lines(send)
        receive_summary = join_lines(recv)

    return {
        "ecu_status_raw": raw_text,
        "ecu_status_std": std_text,
        "send_ecu_summary": send_summary,
        "receive_ecu_summary": receive_summary,
    }


def row_has_signal_definition(ws, row_idx: int, field_map: Dict[str, int]) -> bool:
    signal_name = cell_text(ws.cell(row=row_idx, column=field_map["signal_name"]).value)
    if not is_signal_row_name(signal_name):
        return False

    evidence_fields = ["start_bit", "bit_length", "resolution", "offset", "signal_min", "signal_max"]
    evidence_count = 0
    for code in evidence_fields:
        col = field_map.get(code)
        if col and cell_text(ws.cell(row=row_idx, column=col).value):
            evidence_count += 1

    return evidence_count >= 2


def extract_sheet_records(file_path: Path, ws, arch_version: str, logs: List[str]) -> List[Dict[str, str]]:
    header_row, field_map, missing = find_header_row_and_map(ws)

    if not field_map:
        log(f"[跳过Sheet] 文件={file_path.name} | Sheet={ws.title} | 原因=未通过矩阵核心表头校验", logs)
        return []

    signal_col = field_map["signal_name"]
    ecu_cols = detect_ecu_status_columns(ws, header_row, field_map)

    log(
        f"[识别Sheet] 文件={file_path.name} | Sheet={ws.title} | 表头行={header_row} | "
        f"缺失核心字段={missing if missing else '无'} | ECU状态列数={len(ecu_cols)} | ECU列={[name for _, name in ecu_cols]}",
        logs,
    )

    records: List[Dict[str, str]] = []
    current_msg_info: Dict[str, str] = {}

    for row_idx in range(header_row + 1, ws.max_row + 1):
        signal_name = cell_text(ws.cell(row=row_idx, column=signal_col).value)

        if not row_has_signal_definition(ws, row_idx, field_map):
            for code in ["msg_id", "frame_format", "msg_name", "msg_length", "msg_send_type", "transmitter"]:
                col = field_map.get(code)
                val = cell_text(ws.cell(row=row_idx, column=col).value) if col else ""
                if val:
                    current_msg_info[code] = val
            continue

        rec: Dict[str, str] = {
            "arch_version": arch_version,
            "source_file": file_path.name,
            "source_sheet": ws.title,
            "source_row": str(row_idx),
        }

        for code, _title in OUTPUT_COLUMNS:
            if code in rec:
                continue
            if code in {"ecu_status_raw", "ecu_status_std", "send_ecu_summary", "receive_ecu_summary"}:
                continue

            col = field_map.get(code)
            val = cell_text(ws.cell(row=row_idx, column=col).value) if col else ""
            if not val and code in current_msg_info:
                val = current_msg_info.get(code, "")
            rec[code] = val

        rec["signal_name"] = signal_name

        for code in ["msg_id", "frame_format", "msg_name", "msg_length", "msg_send_type", "transmitter"]:
            if rec.get(code):
                current_msg_info[code] = rec[code]

        rec.update(build_ecu_status_for_row(ws, row_idx, ecu_cols, arch_version))
        records.append(rec)

    log(f"[Sheet完成] {file_path.name} | {ws.title} | 信号数={len(records)}", logs)
    return records


def extract_workbook(file_path: Path, arch_version: str, logs: List[str]) -> List[Dict[str, str]]:
    try:
        wb = load_workbook(file_path, data_only=True, read_only=False)
    except Exception as exc:
        log(f"[读取失败] {file_path.name} | {exc}", logs)
        return []

    all_records: List[Dict[str, str]] = []

    for ws in wb.worksheets:
        try:
            records = extract_sheet_records(file_path, ws, arch_version, logs)
            all_records.extend(records)
        except Exception as exc:
            log(f"[Sheet失败] {file_path.name} | {ws.title} | {exc}", logs)

    wb.close()
    log(f"[文件完成] {file_path.name} | 总信号数={len(all_records)}", logs)
    return all_records


def write_records(records: List[Dict[str, str]], output_file: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [title for _code, title in OUTPUT_COLUMNS]
    codes = [code for code, _title in OUTPUT_COLUMNS]
    ws.append(headers)

    for rec in records:
        ws.append([rec.get(code, "") for code in codes])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    width_map = {
        2: 42, 3: 24, 11: 36, 20: 60, 21: 36,
        22: 95, 23: 95, 24: 45, 25: 45
    }
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = width_map.get(col, 16)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()


def process_dir(input_dir: Path, arch_version: str, output_file: Path, logs: List[str]) -> None:
    if not input_dir.exists():
        raise FileNotFoundError(f"输入目录不存在：{input_dir}")

    files = [p for p in sorted(input_dir.iterdir()) if is_excel_file(p)]
    log(f"\n开始处理 {arch_version}：{input_dir}，Excel数量={len(files)}", logs)

    records: List[Dict[str, str]] = []
    for file_path in files:
        log(f"\n[读取文件] {file_path.name}", logs)
        records.extend(extract_workbook(file_path, arch_version, logs))

    write_records(records, output_file, f"{arch_version}全量信号矩阵清单")
    log(f"[输出完成] {arch_version} | 信号总数={len(records)} | {output_file}", logs)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logs: List[str] = []

    process_dir(INPUT_40_DIR, "4.0", OUTPUT_40_FILE, logs)
    process_dir(INPUT_51_DIR, "5.1", OUTPUT_51_FILE, logs)

    LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
    print(f"\n日志文件：{LOG_FILE}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("脚本运行异常：")
        print(traceback.format_exc())
        sys.exit(1)
