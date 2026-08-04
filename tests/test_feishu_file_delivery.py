from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from core.ai_review import SOURCE_SHEETS
from core.feishu_file_delivery import (
    deliver_task_result_files,
    delivery_file_names,
    register_final_result_files,
)
from core.final_export import FINAL_REVIEW_FILENAME
from core.pipeline import OUTPUT_FILENAMES
from core.review_store import create_task_meta, load_task_meta, update_task_meta


class FakeFileClient:
    def __init__(self, failures: set[str] | None = None) -> None:
        self.failures = set(failures or set())
        self.calls: list[tuple[str, str]] = []

    def send_file(self, file_path: Path, *, chat_id: str) -> dict[str, str]:
        name = Path(file_path).name
        self.calls.append((name, chat_id))
        if name in self.failures:
            raise RuntimeError(f"send denied: {name}")
        index = len(self.calls)
        return {"file_name": name, "file_key": f"file-{index}", "message_id": f"message-{index}"}


def _workbook(path: Path, *, final: bool = False) -> None:
    workbook = Workbook()
    if final:
        workbook.active.title = SOURCE_SHEETS[0]
        workbook.create_sheet(SOURCE_SHEETS[1])
        for sheet_name in SOURCE_SHEETS:
            workbook[sheet_name].append(["4.0信号名", "5.1信号名", "差异点list", "判断结果", "判断来源"])
            workbook[sheet_name].append(["A", "A", "差异", "不同", "人工"])
    else:
        workbook.active.append(["signal"])
        workbook.active.append(["A"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def _task(tmp_path: Path) -> Path:
    tdir = tmp_path / "20260804_120000_test"
    create_task_meta(tdir, tdir.name, status="final_exported")
    output = tdir / "output"
    _workbook(output / OUTPUT_FILENAMES["full_40"])
    _workbook(output / OUTPUT_FILENAMES["full_51"])
    final = output / FINAL_REVIEW_FILENAME
    _workbook(final, final=True)
    register_final_result_files(tdir, final)
    update_task_meta(
        tdir,
        status="final_exported",
        notify_type="feishu_custom_bot",
        review_completed_at="2026-08-04T05:41:03+00:00",
        full_compare_40_parent_url="https://example/pages/通信矩阵-26R1",
        full_compare_51_parent_url="https://example/pages/通信矩阵-26R2",
    )
    return tdir


def test_delivery_names_use_beijing_date_and_each_baseline(tmp_path: Path) -> None:
    tdir = _task(tmp_path)
    names = delivery_file_names(load_task_meta(tdir))
    assert names == {
        "full_40": "20260804_26R1_EEA4.0全量信号矩阵清单.xlsx",
        "full_51": "20260804_26R2_EEA5.1全量信号矩阵清单.xlsx",
        "compare_final": "20260804_EEA4.0和EEA5.1同名信号差异提取.xlsx",
    }


def test_direct_delivery_sends_exactly_three_excel_files_once(tmp_path: Path, monkeypatch) -> None:
    tdir = _task(tmp_path)
    monkeypatch.setenv("FEISHU_RESULT_CHAT_ID", "oc_result_group")
    client = FakeFileClient()

    result = deliver_task_result_files(tdir, client=client)

    assert result["success"] is True
    assert [name for name, _chat in client.calls] == list(delivery_file_names(load_task_meta(tdir)).values())
    assert {chat for _name, chat in client.calls} == {"oc_result_group"}
    meta = load_task_meta(tdir)
    assert meta["status"] == "delivered"
    assert meta["result_delivery_status"] == "delivered"
    assert all(value["status"] == "success" for value in meta["feishu_file_delivery"]["attachments"].values())

    repeated = deliver_task_result_files(tdir, client=client)
    assert repeated["success"] is True
    assert len(client.calls) == 3


def test_partial_failure_retry_only_sends_failed_file(tmp_path: Path, monkeypatch) -> None:
    tdir = _task(tmp_path)
    monkeypatch.setenv("FEISHU_RESULT_CHAT_ID", "oc_result_group")
    names = delivery_file_names(load_task_meta(tdir))
    first = FakeFileClient({names["full_51"]})

    failed = deliver_task_result_files(tdir, client=first)

    assert failed["success"] is False
    assert [entry["status"] for entry in failed["attachments"].values()] == ["success", "failed", "success"]
    retry = FakeFileClient()
    recovered = deliver_task_result_files(tdir, client=retry)
    assert recovered["success"] is True
    assert retry.calls == [(names["full_51"], "oc_result_group")]


def test_missing_group_chat_id_fails_without_sending(tmp_path: Path, monkeypatch) -> None:
    tdir = _task(tmp_path)
    monkeypatch.delenv("FEISHU_RESULT_CHAT_ID", raising=False)
    client = FakeFileClient()

    result = deliver_task_result_files(tdir, client=client)

    assert result["success"] is False
    assert result["failed_stage"] == "target"
    assert "FEISHU_RESULT_CHAT_ID" in result["last_error"]
    assert client.calls == []


def test_final_workbook_with_extra_ai_sheet_is_rejected(tmp_path: Path, monkeypatch) -> None:
    tdir = _task(tmp_path)
    monkeypatch.setenv("FEISHU_RESULT_CHAT_ID", "oc_result_group")
    final = tdir / "output" / FINAL_REVIEW_FILENAME
    workbook = Workbook()
    workbook.active.title = SOURCE_SHEETS[0]
    workbook.create_sheet(SOURCE_SHEETS[1])
    workbook.create_sheet("AI辅助复核与人工审核明细")
    workbook.save(final)
    workbook.close()

    result = deliver_task_result_files(tdir, client=FakeFileClient())

    assert result["success"] is False
    assert result["failed_stage"] == "validate"
