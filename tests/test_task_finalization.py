from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from core.admin_tasks import retry_admin_auto_finalization
from core.ai_review import SOURCE_SHEETS
from core.final_export import FINAL_REVIEW_FILENAME, SHEET_RULES
from core.notification_router import scan_custom_notifications
from core.pipeline import OUTPUT_FILENAMES
from core.review_store import (
    create_task_meta,
    init_review_state,
    load_task_meta,
    save_review_state,
    update_task_meta,
)
from core.task_finalization import auto_finalize_if_no_pending


def _item() -> dict:
    return {
        "item_id": "item-1",
        "source_sheet": "完全同名匹配对比结果",
        "signal_40": "SignalA",
        "signal_51": "SignalA",
        "field_diffs": [
            {
                "diff_field": "信号值描述",
                "field_type": "text",
                "value_40": "0x0: Off",
                "value_51": "0x0: Not active",
            }
        ],
        "signal_ai_judgement": "疑似可忽略",
        "signal_ai_reason": "表达方式不同",
    }


def _task(tmp_path: Path, *, reviewed: bool) -> Path:
    tdir = tmp_path / ("covered" if reviewed else "pending")
    review_dir = tdir / "review"
    output_dir = tdir / "output"
    output_dir.mkdir(parents=True)
    create_task_meta(tdir, tdir.name, status="awaiting_review")
    item = _item()
    compare = Workbook()
    compare.active.title = SOURCE_SHEETS[0]
    compare.create_sheet(SOURCE_SHEETS[1])
    for sheet_name in SOURCE_SHEETS:
        headers = ["4.0信号名", "5.1信号名"]
        if sheet_name == SOURCE_SHEETS[1]:
            headers.append("去前缀后匹配名")
        headers.append("差异点list")
        for field in (
            "信号长度", "精度", "偏移量", "物理最小值", "物理最大值", "单位", "信号值描述",
            "信号来源文件", "ECU收发状态_原始", "ECU收发状态_标准化", "发送ECU汇总", "接收ECU汇总",
        ):
            headers.extend((f"4.0_{field}", f"5.1_{field}"))
        compare[sheet_name].append(headers)
    exact = compare[SOURCE_SHEETS[0]]
    row = [""] * exact.max_column
    headers = [cell.value for cell in exact[1]]
    row[headers.index("4.0信号名")] = item["signal_40"]
    row[headers.index("5.1信号名")] = item["signal_51"]
    row[headers.index("差异点list")] = "信号值描述差异"
    exact.append(row)
    compare.save(output_dir / OUTPUT_FILENAMES["compare"])
    compare.close()
    (review_dir / "review_items.json").parent.mkdir(parents=True, exist_ok=True)
    (review_dir / "review_items.json").write_text(
        json.dumps([item], ensure_ascii=False), encoding="utf-8"
    )
    state = init_review_state(review_dir, tdir.name, [item])
    if reviewed:
        field = state["items"]["item-1"]["field_reviews"]["信号值描述"]
        field.update(
            result="same",
            reviewed=True,
            decision_source="history_manual",
            reviewer="history",
            reviewed_at="2026-08-04T02:00:00+00:00",
        )
        save_review_state(review_dir, state)
    update_task_meta(
        tdir,
        status="awaiting_review",
        notify_type="feishu_custom_bot",
        pending_manual_count=0 if reviewed else 1,
        history_reused_count=1 if reviewed else 0,
    )
    return tdir


def test_all_history_covered_task_skips_review_and_generates_final_result(
    tmp_path: Path, monkeypatch
) -> None:
    tdir = _task(tmp_path, reviewed=True)
    deliveries: list[Path] = []
    monkeypatch.setattr(
        "core.feishu_sheet_delivery.deliver_task_result_sheets",
        lambda task_dir, **_kwargs: deliveries.append(Path(task_dir)) or {"success": True},
    )

    result = auto_finalize_if_no_pending(tdir)

    assert result["success"] is True
    final_path = tdir / "output" / FINAL_REVIEW_FILENAME
    assert final_path.is_file() and final_path.stat().st_size > 0
    workbook = load_workbook(final_path, read_only=True)
    try:
        assert workbook.sheetnames == SHEET_RULES
        assert workbook[SOURCE_SHEETS[0]].max_row == 1
        headers = [cell.value for cell in workbook[SOURCE_SHEETS[0]][1]]
        assert "判断结果" in headers and "判断来源" in headers
    finally:
        workbook.close()
    meta = load_task_meta(tdir)
    assert meta["status"] == "final_exported"
    assert meta["review_completed"] is True
    assert meta["auto_finalization_status"] == "success"
    assert meta["auto_finalized_without_manual"] is True
    assert meta["auto_finalized_by_history"] is True
    assert meta["pending_manual_count"] == 0
    assert meta["result_url"]
    assert deliveries == [tdir.resolve()]

    repeated = auto_finalize_if_no_pending(tdir)
    assert repeated["success"] is True and repeated["already_finalized"] is True
    assert deliveries == [tdir.resolve()]


def test_pending_task_stays_on_manual_review_path(tmp_path: Path) -> None:
    tdir = _task(tmp_path, reviewed=False)

    result = auto_finalize_if_no_pending(tdir)

    assert result == {
        "success": False,
        "skipped": True,
        "reason": "manual_review_required",
        "pending_manual": 1,
    }
    assert load_task_meta(tdir)["status"] == "awaiting_review"
    assert not (tdir / "output" / FINAL_REVIEW_FILENAME).exists()


def test_existing_zero_pending_task_is_recovered_by_notification_scan(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("TASK_ROOT_DIR", str(tmp_path))
    tdir = _task(tmp_path, reviewed=True)
    calls: list[tuple[Path, bool]] = []
    monkeypatch.setattr(
        "core.task_finalization.auto_finalize_if_no_pending",
        lambda task_dir, notify=True, **_kwargs: calls.append((Path(task_dir), notify))
        or {"success": True},
    )

    scan_custom_notifications()

    assert calls == [(tdir, True)]


def test_admin_can_force_retry_zero_pending_finalization(
    tmp_path: Path, monkeypatch
) -> None:
    monkeypatch.setenv("TASK_ROOT_DIR", str(tmp_path))
    tdir = _task(tmp_path, reviewed=True)
    calls: list[tuple[Path, bool, bool]] = []
    monkeypatch.setattr(
        "core.task_finalization.auto_finalize_if_no_pending",
        lambda task_dir, notify=True, force=False: calls.append(
            (Path(task_dir), notify, force)
        )
        or {"success": True},
    )

    assert retry_admin_auto_finalization(tdir.name)["success"] is True
    assert calls == [(tdir.resolve(), True, True)]


def test_process_claim_prevents_duplicate_final_generation(
    tmp_path: Path, monkeypatch
) -> None:
    tdir = _task(tmp_path, reviewed=True)
    claim = tdir / "bot" / "auto_finalization.lock"
    claim.parent.mkdir(parents=True, exist_ok=True)
    claim.write_text("busy", encoding="utf-8")
    monkeypatch.setenv("AUTO_FINALIZE_LOCK_STALE_SECONDS", "3600")

    result = auto_finalize_if_no_pending(tdir)

    assert result == {"success": False, "running": True, "reason": "already_running"}
    assert not (tdir / "output" / FINAL_REVIEW_FILENAME).exists()
