from __future__ import annotations

import json
from pathlib import Path
import sys

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook

from core.ai_review import SOURCE_SHEETS
from core.final_export import export_final_review_result
from core.review_history import (
    admin_set_history_enabled,
    admin_update_history_decision,
    export_history_csv,
    history_counts,
    history_database_path,
    history_decision_events,
    history_summary,
    list_history_decisions,
)
from core.review_store import acquire_review_lock, compute_review_stats, create_task_meta, init_review_state
from core.review_table import save_dirty_reviews


def _item(item_id: str, *, description_51: str = "Inactive", with_unit: bool = False) -> dict:
    diffs = [
        {"diff_field": "信号值描述", "value_40": "Active", "value_51": description_51, "field_type": "text"},
    ]
    if with_unit:
        diffs.append({"diff_field": "单位", "value_40": "Nm", "value_51": "N·m", "field_type": "text"})
    return {
        "item_id": item_id,
        "source_sheet": "完全同名匹配对比结果",
        "signal_40": "ExampleSignal",
        "signal_51": "ExampleSignal",
        "field_diffs": diffs,
        "signal_ai_judgement": "疑似可忽略",
    }


def _compare_file(path: Path, item: dict) -> Path:
    workbook = Workbook()
    workbook.active.title = SOURCE_SHEETS[0]
    workbook.create_sheet(SOURCE_SHEETS[1])
    for sheet_name in SOURCE_SHEETS:
        workbook[sheet_name].append(["4.0信号名", "5.1信号名", "差异点list"])
    workbook[SOURCE_SHEETS[0]].append([item["signal_40"], item["signal_51"], "差异"])
    workbook.save(path)
    workbook.close()
    return path


def _save_task_decisions(task_dir: Path, task_id: str, item: dict, results: dict[str, str]) -> None:
    review_dir = task_dir / "review"
    review_dir.mkdir(parents=True, exist_ok=True)
    (review_dir / "review_items.json").write_text(json.dumps([item], ensure_ascii=False), encoding="utf-8")
    state = init_review_state(review_dir, task_id, [item])
    acquire_review_lock(task_dir, "reviewer-1")
    drafts = {
        f"{item['item_id']}::{field}": {"item_id": item["item_id"], "field_key": field, "result": result}
        for field, result in results.items()
    }
    save_dirty_reviews(
        review_dir, task_id, drafts, set(drafts),
        base_revision=int(state["revision"]), session_id="reviewer-1",
    )


def test_exact_history_is_applied_to_new_task(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "different"})

    second = tmp_path / "second"
    create_task_meta(second, "second")
    item = _item("new")
    state = init_review_state(second / "review", "second", [item])
    review = state["items"]["new"]["field_reviews"]["信号值描述"]
    assert review["result"] == "different"
    assert review["reviewed"] is True
    assert review["decision_source"] == "history_manual"
    assert review["history_task_id"] == "first"
    assert compute_review_stats([item], state)["history_reused"] == 1

    items_path = second / "review" / "review_items.json"
    items_path.write_text(json.dumps([item], ensure_ascii=False), encoding="utf-8")
    output = tmp_path / "history-result.xlsx"
    export_final_review_result(
        items_path,
        second / "review" / "review_state.json",
        output,
        compare_file_path=_compare_file(tmp_path / "history-source.xlsx", item),
    )
    workbook = load_workbook(output, read_only=True)
    try:
        headers = [cell.value for cell in workbook[SOURCE_SHEETS[0]][1]]
        assert workbook[SOURCE_SHEETS[0]].cell(2, headers.index("判断来源") + 1).value == "人工"
    finally:
        workbook.close()


def test_changed_field_value_does_not_reuse_history(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "same"})

    second = tmp_path / "second"
    create_task_meta(second, "second")
    changed = _item("new", description_51="Disabled")
    state = init_review_state(second / "review", "second", [changed])
    review = state["items"]["new"]["field_reviews"]["信号值描述"]
    assert review["reviewed"] is False
    assert review["result"] == ""


def test_description_and_unit_history_are_reused_independently(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "same"})

    second = tmp_path / "second"
    create_task_meta(second, "second")
    combined = _item("new", with_unit=True)
    state = init_review_state(second / "review", "second", [combined])
    fields = state["items"]["new"]["field_reviews"]
    assert fields["信号值描述"]["decision_source"] == "history_manual"
    assert fields["单位"]["reviewed"] is False


def test_manual_override_updates_history_and_keeps_audit_events(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "same"})
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "different"})
    assert history_counts(db_path=history_database_path(first / "review")) == {"decisions": 1, "events": 2}

    second = tmp_path / "second"
    create_task_meta(second, "second")
    state = init_review_state(second / "review", "second", [_item("new")])
    assert state["items"]["new"]["field_reviews"]["信号值描述"]["result"] == "different"


def test_admin_can_search_correct_disable_and_restore_history(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "same"})
    db_path = history_database_path(first / "review")

    summary = history_summary(db_path=db_path)
    assert summary["decisions"] == 1
    assert summary["enabled"] == 1
    assert summary["descriptions"] == 1
    rows = list_history_decisions(db_path=db_path, search="ExampleSignal", result="same", enabled=True)
    assert len(rows) == 1
    fingerprint = rows[0]["fingerprint"]

    updated = admin_update_history_decision(
        db_path=db_path, fingerprint=fingerprint, result="different",
        actor="admin_web", reason="业务规则纠正",
    )
    assert updated["result"] == "different"
    assert history_decision_events(db_path=db_path, fingerprint=fingerprint)[0]["action"] == "admin_correct"

    admin_set_history_enabled(
        db_path=db_path, fingerprint=fingerprint, enabled=False,
        actor="admin_web", reason="等待复核",
    )
    assert history_summary(db_path=db_path)["disabled"] == 1
    second = tmp_path / "second"
    create_task_meta(second, "second")
    state = init_review_state(second / "review", "second", [_item("new")])
    assert state["items"]["new"]["field_reviews"]["信号值描述"]["reviewed"] is False

    admin_set_history_enabled(
        db_path=db_path, fingerprint=fingerprint, enabled=True,
        actor="admin_web", reason="复核通过",
    )
    third = tmp_path / "third"
    create_task_meta(third, "third")
    state = init_review_state(third / "review", "third", [_item("newer")])
    assert state["items"]["newer"]["field_reviews"]["信号值描述"]["result"] == "different"


def test_admin_history_requires_reason_and_exports_audit(tmp_path: Path) -> None:
    first = tmp_path / "first"
    create_task_meta(first, "first")
    _save_task_decisions(first, "first", _item("old"), {"信号值描述": "same"})
    db_path = history_database_path(first / "review")
    fingerprint = list_history_decisions(db_path=db_path)[0]["fingerprint"]

    with pytest.raises(ValueError, match="必须填写原因"):
        admin_update_history_decision(
            db_path=db_path, fingerprint=fingerprint, result="different", actor="admin_web", reason="",
        )
    with pytest.raises(ValueError, match="必须填写原因"):
        admin_set_history_enabled(
            db_path=db_path, fingerprint=fingerprint, enabled=False, actor="admin_web", reason="",
        )
    assert "ExampleSignal" in export_history_csv(db_path=db_path).decode("utf-8-sig")
    assert "confirmed_at" in export_history_csv(db_path=db_path, events=True).decode("utf-8-sig")
