from pathlib import Path
import json
import sys

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from core.final_export import (
    DESCRIPTION_ANCHOR_HEADER,
    FINAL_COLUMN_WIDTHS,
    FINAL_RESULT_HEADERS,
    SHEET_RULES,
    SOURCE_FILE_HEADERS,
    SUMMARY_HEADERS,
    export_final_review_result,
)
from core.ai_review import AI_REVIEW_SHEET, SOURCE_SHEETS, is_text_only_ai_candidate
from core.review_store import acquire_review_lock, create_task_meta, init_review_state, save_review_state, update_review_field


def _text_item() -> dict:
    return {
        "item_id": "signal-1", "source_sheet": "完全同名匹配对比结果", "signal_40": "S40", "signal_51": "S51",
        "field_diffs": [
            {"diff_field": "信号值描述", "value_40": "Off", "value_51": "Disable", "field_type": "text"},
            {"diff_field": "单位", "value_40": "Nm", "value_51": "N·m", "field_type": "text"},
        ],
        "signal_ai_judgement": "疑似可忽略", "signal_ai_reason": "文本可能等价",
    }


def _numeric_item() -> dict:
    return {
        "item_id": "signal-2", "source_sheet": "vcu-hcu 同名匹配", "signal_40": "N40", "signal_51": "N51",
        "field_diffs": [
            {"diff_field": "信号值描述", "value_40": "Off", "value_51": "Disable", "field_type": "text"},
            {"diff_field": "信号长度", "value_40": "8", "value_51": "12", "field_type": "numeric"},
        ],
        "signal_ai_judgement": "真实差异", "signal_ai_reason": "包含数值差异",
    }


def _same_item() -> dict:
    return {
        "item_id": "signal-3", "source_sheet": "完全同名匹配对比结果", "signal_40": "Same40", "signal_51": "Same51",
        "field_diffs": [
            {"diff_field": "单位", "value_40": "km/h", "value_51": "kmh", "field_type": "text"},
        ],
        "signal_ai_judgement": "疑似可忽略", "signal_ai_reason": "单位表达等价",
    }


def _compare_workbook(path: Path, items: list[dict]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    compare_fields = ["信号长度", "精度", "偏移量", "物理最小值", "物理最大值", "单位", "信号值描述"]
    extra_fields = ["信号来源文件", "ECU收发状态_原始", "ECU收发状态_标准化", "发送ECU汇总", "接收ECU汇总"]
    for sheet_name in SOURCE_SHEETS:
        ws = workbook.create_sheet(sheet_name)
        headers = ["4.0信号名", "5.1信号名"]
        if sheet_name == SOURCE_SHEETS[1]:
            headers.append("去前缀后匹配名")
        headers.append("差异点list")
        for field in (*compare_fields, *extra_fields):
            headers.extend((f"4.0_{field}", f"5.1_{field}"))
        ws.append(headers)
        for column, header in enumerate(headers, start=1):
            if "描述" in header or "差异点" in header or "ECU" in header or "来源" in header:
                ws.column_dimensions[ws.cell(1, column).column_letter].width = 95
            elif "信号名" in header:
                ws.column_dimensions[ws.cell(1, column).column_letter].width = 36
            else:
                ws.column_dimensions[ws.cell(1, column).column_letter].width = 16
        for item in items:
            if item["source_sheet"] == sheet_name:
                values = {header: f"{item['item_id']}-{header}" for header in headers}
                values["4.0信号名"] = item["signal_40"]
                values["5.1信号名"] = item["signal_51"]
                values["差异点list"] = "差异内容"
                ws.append([values[header] for header in headers])
    workbook.create_sheet(AI_REVIEW_SHEET).append(["内部AI明细"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def test_field_review_updates_existing_value_and_numeric_is_system_different(tmp_path: Path) -> None:
    task_dir = tmp_path / "task"
    review_dir = task_dir / "review"
    create_task_meta(task_dir, "task", status="reviewing")
    state = init_review_state(review_dir, "task", [_numeric_item()])
    fields = state["items"]["signal-2"]["field_reviews"]
    assert fields["信号长度"]["result"] == "different"
    assert fields["信号长度"]["decision_source"] == "system_default"
    assert fields["信号值描述"]["result"] == "different"
    assert fields["信号值描述"]["decision_source"] == "system_default"

    # Repeated manual confirmation updates the same field instead of adding a record.
    state = init_review_state(review_dir, "task", [_text_item()], overwrite=True)
    acquire_review_lock(task_dir, "session")
    state = update_review_field(review_dir, "task", "signal-1", "信号值描述", "same", base_revision=0, session_id="session")
    state = update_review_field(review_dir, "task", "signal-1", "信号值描述", "different", base_revision=1, session_id="session")
    assert state["items"]["signal-1"]["field_reviews"]["信号值描述"]["result"] == "different"
    assert len(state["items"]["signal-1"]["field_reviews"]) == 2


def test_final_export_keeps_two_source_sheets_and_adds_signal_decisions(tmp_path: Path) -> None:
    task_dir = tmp_path / "task"
    review_dir = task_dir / "review"
    create_task_meta(task_dir, "task", status="reviewing")
    items = [_text_item(), _numeric_item(), _same_item()]
    init_review_state(review_dir, "task", items)
    acquire_review_lock(task_dir, "session")
    update_review_field(review_dir, "task", "signal-1", "信号值描述", "different", base_revision=0, session_id="session")
    update_review_field(review_dir, "task", "signal-1", "单位", "same", base_revision=1, session_id="session")
    update_review_field(review_dir, "task", "signal-3", "单位", "same", base_revision=2, session_id="session")
    items_path = review_dir / "review_items.json"
    items_path.parent.mkdir(parents=True, exist_ok=True)
    items_path.write_text(json.dumps(items, ensure_ascii=False), encoding="utf-8")
    output = tmp_path / "final.xlsx"
    compare = tmp_path / "compare.xlsx"
    _compare_workbook(compare, items)
    stats = export_final_review_result(items_path, review_dir / "review_state.json", output, compare_file_path=compare)
    assert stats == {"判断结果-相同": 1, "判断结果-不同": 2, "判断来源-人工": 2, "判断来源-系统": 1, "审核信号总数": 3}
    wb = load_workbook(output)
    try:
        assert wb.sheetnames == SHEET_RULES
        assert AI_REVIEW_SHEET not in wb.sheetnames
        for sheet_name, expected_source in ((SOURCE_SHEETS[0], "人工"), (SOURCE_SHEETS[1], "系统")):
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            diff_index = headers.index("差异点list")
            assert tuple(headers[diff_index + 1:diff_index + 3]) == FINAL_RESULT_HEADERS
            assert ws.cell(2, diff_index + 2).value == "不同"
            assert ws.cell(2, diff_index + 3).value == expected_source
            description_index = headers.index(DESCRIPTION_ANCHOR_HEADER)
            assert tuple(headers[description_index + 1:description_index + 5]) == SUMMARY_HEADERS
            assert tuple(headers[-2:]) == SOURCE_FILE_HEADERS
            assert "相同" not in {
                ws.cell(row, headers.index("判断结果") + 1).value
                for row in range(2, ws.max_row + 1)
            }
            assert "Same40" not in {
                ws.cell(row, headers.index("4.0信号名") + 1).value
                for row in range(2, ws.max_row + 1)
            }
            header_columns = {cell.value: cell.column_letter for cell in ws[1]}
            assert {
                header: ws.column_dimensions[header_columns[header]].width
                for header in FINAL_COLUMN_WIDTHS if header in header_columns
            } == {header: width for header, width in FINAL_COLUMN_WIDTHS.items() if header in header_columns}
    finally:
        wb.close()


def test_ai_candidates_are_strictly_description_and_unit_only() -> None:
    assert is_text_only_ai_candidate(_text_item()) is True
    assert is_text_only_ai_candidate(_numeric_item()) is False
    unknown = _text_item()
    unknown["field_diffs"].append({"diff_field": "未解析", "value_40": "", "value_51": "", "field_type": "unknown"})
    assert is_text_only_ai_candidate(unknown) is False


def test_existing_mixed_numeric_state_is_reclassified_without_rerunning_task(tmp_path: Path) -> None:
    review_dir = tmp_path / "task" / "review"
    save_review_state(review_dir, {
        "task_id": "task", "revision": 0, "items": {"signal-2": {"field_reviews": {
            "信号值描述": {"diff_field": "信号值描述", "result": "", "reviewed": False, "decision_source": ""},
            "信号长度": {"diff_field": "信号长度", "result": "different", "reviewed": True, "decision_source": "system_default"},
        }}}
    })
    state = init_review_state(review_dir, "task", [_numeric_item()])
    description = state["items"]["signal-2"]["field_reviews"]["信号值描述"]
    assert description["result"] == "different"
    assert description["decision_source"] == "system_default"
