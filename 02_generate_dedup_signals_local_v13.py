# -*- coding: utf-8 -*-
"""
02_generate_dedup_signals_local_v13.py

基于文件一、文件二生成文件三、文件四。

本版口径：
1. 按信号名去重；
2. 信号定义字段保留首次出现记录；
3. 信号来源文件汇总所有出现文件；
4. ECU 收发状态保留并汇总：
   - 4.0 在全量阶段已经是主 ECU:状态(后续 ECU:状态...) 的括号格式；
   - 去重阶段保持该格式，不拆散；
   - 标准化状态使用小写 r/s。
"""

from __future__ import annotations

import sys
import traceback
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from common_matrix_utils_local_v13 import (
    cell_text,
    norm_header,
    join_lines,
    join_unique_status_texts,
)


BASE_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = BASE_DIR / "output"

INPUT_40_FILE = OUTPUT_DIR / "26R1 4.0全量信号矩阵清单.xlsx"
INPUT_51_FILE = OUTPUT_DIR / "26R2 5.1全量信号矩阵清单.xlsx"

OUTPUT_40_FILE = OUTPUT_DIR / "26R1 4.0全量信号-同名去重后.xlsx"
OUTPUT_51_FILE = OUTPUT_DIR / "26R2 5.1全量信号-同名去重后.xlsx"
LOG_FILE = OUTPUT_DIR / "02_generate_dedup_signals_local_v13_log.txt"


BUSINESS_FIELDS = [
    ("signal_name", "信号名称"),
    ("bit_length", "信号长度"),
    ("resolution", "精度"),
    ("offset", "偏移量"),
    ("signal_min", "物理最小值"),
    ("signal_max", "物理最大值"),
    ("unit", "单位"),
    ("signal_value_description", "信号值描述"),
]

EXTRA_FIELDS = [
    ("source_files", "信号来源文件"),
    ("ecu_status_raw", "ECU收发状态_原始"),
    ("ecu_status_std", "ECU收发状态_标准化"),
    ("send_ecu_summary", "发送ECU汇总"),
    ("receive_ecu_summary", "接收ECU汇总"),
]

OUTPUT_FIELDS = BUSINESS_FIELDS + EXTRA_FIELDS

COLUMN_CANDIDATES = {
    "signal_name": ["信号名称", "Signal Name", "Signal Name 信号名称", "信号名"],
    "bit_length": ["信号长度", "Bit Length", "Signal Size", "Signal Length"],
    "resolution": ["精度", "Resolution", "Factor", "Scale"],
    "offset": ["偏移量", "Offset"],
    "signal_min": ["物理最小值", "Signal Min", "Minimum", "Min"],
    "signal_max": ["物理最大值", "Signal Max", "Maximum", "Max"],
    "unit": ["单位", "Unit"],
    "signal_value_description": ["信号值描述", "Signal Value Description", "Value Description"],
    "source_file": ["来源文件"],
    "ecu_status_raw": ["ECU收发状态_原始", "ECU接收和发送状态"],
    "ecu_status_std": ["ECU收发状态_标准化"],
    "send_ecu_summary": ["发送ECU汇总"],
    "receive_ecu_summary": ["接收ECU汇总"],
}


def match_col(value: Any, names: List[str]) -> bool:
    raw = cell_text(value)
    compact = norm_header(raw)
    if not raw:
        return False
    for name in names:
        n = norm_header(name)
        if compact == n or n in compact:
            return True
    return False


def find_columns(ws) -> Tuple[int, Dict[str, int]]:
    best_row = 1
    best_map: Dict[str, int] = {}
    best_score = -1

    for row_idx in range(1, min(ws.max_row or 0, 10) + 1):
        cmap: Dict[str, int] = {}
        for col_idx in range(1, (ws.max_column or 0) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            for code, names in COLUMN_CANDIDATES.items():
                if code not in cmap and match_col(v, names):
                    cmap[code] = col_idx
                    break
        if "signal_name" in cmap and len(cmap) > best_score:
            best_row = row_idx
            best_map = cmap
            best_score = len(cmap)

    if "signal_name" not in best_map:
        raise ValueError(f"未找到信号名称列：{ws.title}")

    return best_row, best_map


def read_records(input_file: Path) -> List[Dict[str, str]]:
    wb = load_workbook(input_file, data_only=True, read_only=False)
    ws = wb.active
    header_row, cmap = find_columns(ws)

    records: List[Dict[str, str]] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        name = cell_text(ws.cell(row=row_idx, column=cmap["signal_name"]).value)
        if not name:
            continue

        rec: Dict[str, str] = {}
        for code, _title in BUSINESS_FIELDS:
            col = cmap.get(code)
            rec[code] = cell_text(ws.cell(row=row_idx, column=col).value) if col else ""

        for code in ["source_file", "ecu_status_raw", "ecu_status_std", "send_ecu_summary", "receive_ecu_summary"]:
            col = cmap.get(code)
            rec[code] = cell_text(ws.cell(row=row_idx, column=col).value) if col else ""

        records.append(rec)

    wb.close()
    return records


def dedup(records: List[Dict[str, str]]) -> List[Dict[str, str]]:
    groups: Dict[str, List[Dict[str, str]]] = {}
    order: List[str] = []

    for rec in records:
        name = cell_text(rec.get("signal_name"))
        if not name:
            continue
        if name not in groups:
            groups[name] = []
            order.append(name)
        groups[name].append(rec)

    out: List[Dict[str, str]] = []

    for name in order:
        group = groups[name]
        first = group[0]

        new: Dict[str, str] = {}
        for code, _title in BUSINESS_FIELDS:
            new[code] = first.get(code, "")

        new["source_files"] = join_lines([r.get("source_file", "") for r in group])

        # 保持全量阶段的格式，尤其是 4.0 的括号格式，不拆散。
        # 注意：发送ECU汇总、接收ECU汇总也直接聚合全量文件中的格式，
        # 这样 4.0 会继续保持 EMS_PHEV(HighRegulationArea,LowRegulationArea) 这种括号格式。
        new["ecu_status_raw"] = join_unique_status_texts([r.get("ecu_status_raw", "") for r in group])
        new["ecu_status_std"] = join_unique_status_texts([r.get("ecu_status_std", "") for r in group])
        new["send_ecu_summary"] = join_unique_status_texts([r.get("send_ecu_summary", "") for r in group])
        new["receive_ecu_summary"] = join_unique_status_texts([r.get("receive_ecu_summary", "") for r in group])

        out.append(new)

    return out


def write_output(records: List[Dict[str, str]], output_file: Path, sheet_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = [title for _code, title in OUTPUT_FIELDS]
    codes = [code for code, _title in OUTPUT_FIELDS]
    ws.append(headers)

    for rec in records:
        ws.append([rec.get(code, "") for code in codes])

    header_fill = PatternFill("solid", fgColor="5B9BD5")
    header_font = Font(bold=True, color="FFFFFF")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    widths = {
        1: 36, 8: 70, 9: 50, 10: 95, 11: 95, 12: 48, 13: 48
    }
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 16)

    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    wb.close()


def process(input_file: Path, output_file: Path, sheet_name: str, logs: List[str]) -> None:
    records = read_records(input_file)
    deduped = dedup(records)
    write_output(deduped, output_file, sheet_name)

    msg = f"{input_file.name}：原始行数={len(records)}，去重后={len(deduped)}，输出={output_file}"
    print(msg, flush=True)
    logs.append(msg)


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    logs: List[str] = []

    process(INPUT_40_FILE, OUTPUT_40_FILE, "4.0同名去重后", logs)
    process(INPUT_51_FILE, OUTPUT_51_FILE, "5.1同名去重后", logs)

    LOG_FILE.write_text("\n".join(logs), encoding="utf-8")
    print(f"日志文件：{LOG_FILE}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        print("脚本运行异常：")
        print(traceback.format_exc())
        sys.exit(1)
